#!/usr/bin/env python3
"""
Rigid Plot_3D Template System

Creates protected, format-compliant templates for Plot_3D that ensure:
- Exact column structure required for K-means clustering
- Proper format for ΔE calculations 
- Compatibility with "Refresh Data" functionality
- Prevention of user format corruption
- Professional worksheet protection

The templates maintain strict format compliance while allowing data entry
only in designated areas.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
import logging
import os
from datetime import datetime

logger = logging.getLogger(__name__)


class RigidPlot3DTemplate:
    """Creates rigid, protected Plot_3D templates with exact format compliance."""
    
    # Exact Plot_3D column structure - DO NOT MODIFY
    PLOT3D_COLUMNS = [
        'Xnorm', 'Ynorm', 'Znorm', 'DataID', 'Cluster', 
        '∆E', 'Marker', 'Color', 'Centroid_X', 'Centroid_Y', 
        'Centroid_Z', 'Sphere', 'Radius'
    ]
    
    # Valid values for data validation (from Plot_3D requirements)
    VALID_MARKERS = ['.', 'o', '*', '^', '<', '>', 'v', 's', 'D', '+', 'x']
    VALID_COLORS = [
        'red', 'blue', 'green', 'orange', 'purple', 'yellow', 
        'cyan', 'magenta', 'brown', 'pink', 'lime', 'navy', 'teal', 'gray'
    ]
    VALID_SPHERES = [
        'red', 'green', 'blue', 'yellow', 'cyan', 'magenta', 
        'orange', 'purple', 'brown', 'pink', 'lime', 'navy', 'teal', 'gray'
    ]
    
    # Protection settings
    SHEET_PASSWORD = "Plot3D_Protected"  # Can be changed if needed
    
    def __init__(self):
        """Initialize the rigid template creator."""
        self.workbook = None
        self.worksheet = None
        
    def create_rigid_template(self, file_path: str, sample_set_name: str = "Plot3D_Analysis") -> bool:
        """
        Create a rigid, protected Plot_3D template.
        
        Args:
            file_path: Path where to save the template
            sample_set_name: Name for the sample set (used in metadata)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Create new workbook
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Plot3D_Data"
            
            # Set up the rigid structure
            self._create_metadata_section(sample_set_name)
            self._setup_rigid_headers()
            self._apply_rigid_formatting()
            self._setup_rigid_validation()
            self._apply_sheet_protection()
            self._add_instructions()
            
            # Save the protected template
            self.workbook.save(file_path)
            logger.info(f"Created rigid Plot_3D template: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating rigid template: {e}")
            return False
    
    def _create_metadata_section(self, sample_set_name: str):
        """Create protected metadata section (rows 1-7)."""
        # Template identification
        self.worksheet['A1'] = "Plot_3D Data Template"
        self.worksheet['A1'].font = Font(bold=True, size=14)
        
        self.worksheet['A2'] = f"Sample Set: {sample_set_name}"
        self.worksheet['A2'].font = Font(bold=True)
        
        self.worksheet['A3'] = f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Format requirements notice
        self.worksheet['A5'] = "IMPORTANT: This template format is required for Plot_3D functionality"
        self.worksheet['A5'].font = Font(bold=True, color="FF0000")  # Red text
        
        self.worksheet['A6'] = "• K-means clustering requires exact column structure"
        self.worksheet['A7'] = "• ΔE calculations depend on proper data format"
        
        # Data entry instructions
        self.worksheet['G1'] = "DATA ENTRY AREA:"
        self.worksheet['G1'].font = Font(bold=True, color="0000FF")  # Blue text
        
        self.worksheet['G2'] = "• Enter data starting at row 9"
        self.worksheet['G3'] = "• Use dropdown lists in gray columns"
        self.worksheet['G4'] = "• DO NOT modify column headers"
        self.worksheet['G5'] = "• Save file after changes for 'Refresh Data'"
    
    def _setup_rigid_headers(self):
        """Set up protected column headers in row 8."""
        header_row = 8
        
        for col_idx, header in enumerate(self.PLOT3D_COLUMNS, 1):
            cell = self.worksheet.cell(row=header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Special formatting for key columns
            if header in ['Xnorm', 'Ynorm', 'Znorm']:
                cell.fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')  # Light red
            elif header == 'DataID':
                cell.fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')  # Light blue
            elif header in ['Cluster', '∆E']:
                cell.fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')  # Light green
    
    def _apply_rigid_formatting(self):
        """Apply rigid formatting with clear visual indicators."""
        # Protected metadata area (rows 1-7) - Light yellow background
        metadata_fill = PatternFill(start_color='FFFACD', end_color='FFFACD', fill_type='solid')
        for row in range(1, 8):
            for col in range(1, 14):  # All columns
                cell = self.worksheet.cell(row=row, column=col)
                cell.fill = metadata_fill
        
        # Header row (row 8) - Already formatted in _setup_rigid_headers
        
        # Data validation areas - Gray background for dropdown columns
        gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        # Marker column (G) - rows 9-200
        for row in range(9, 201):
            cell = self.worksheet.cell(row=row, column=7)  # Column G
            cell.fill = gray_fill
            
        # Color column (H) - rows 9-200  
        for row in range(9, 201):
            cell = self.worksheet.cell(row=row, column=8)  # Column H
            cell.fill = gray_fill
            
        # Sphere column (L) - rows 9-200
        for row in range(9, 201):
            cell = self.worksheet.cell(row=row, column=12)  # Column L
            cell.fill = gray_fill
        
        # Add borders to data area
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to header and first few data rows
        for row in range(8, 12):  # Header + first 3 data rows
            for col in range(1, 14):
                self.worksheet.cell(row=row, column=col).border = thin_border
    
    def _setup_rigid_validation(self):
        """Set up rigid data validation that cannot be modified by users."""
        # Marker validation (Column G, rows 9-200)
        marker_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(self.VALID_MARKERS)}"',
            allow_blank=True,
            showDropDown=True,
            showErrorMessage=True
        )
        marker_validation.error = "Invalid marker! Use dropdown to select valid marker type."
        marker_validation.errorTitle = "Plot_3D Format Error"
        marker_validation.prompt = "Select marker type for 3D visualization"
        marker_validation.promptTitle = "Marker Selection"
        
        self.worksheet.add_data_validation(marker_validation)
        marker_validation.add("G9:G200")
        
        # Color validation (Column H, rows 9-200)
        color_validation = DataValidation(
            type="list", 
            formula1=f'"{",".join(self.VALID_COLORS)}"',
            allow_blank=True,
            showDropDown=True,
            showErrorMessage=True
        )
        color_validation.error = "Invalid color! Use dropdown to select valid color name."
        color_validation.errorTitle = "Plot_3D Format Error"
        color_validation.prompt = "Select color for 3D point visualization"
        color_validation.promptTitle = "Color Selection"
        
        self.worksheet.add_data_validation(color_validation)
        color_validation.add("H9:H200")
        
        # Sphere validation (Column L, rows 9-200)
        sphere_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(self.VALID_SPHERES)}"',
            allow_blank=True,
            showDropDown=True,
            showErrorMessage=True
        )
        sphere_validation.error = "Invalid sphere color! Use dropdown to select valid color."
        sphere_validation.errorTitle = "Plot_3D Format Error"
        sphere_validation.prompt = "Select sphere color for clustering visualization"
        sphere_validation.promptTitle = "Sphere Color Selection"
        
        self.worksheet.add_data_validation(sphere_validation)
        sphere_validation.add("L9:L200")
    
    def _apply_sheet_protection(self):
        """Apply sheet protection while allowing data entry in specific cells."""
        # Configure protection settings
        protection = SheetProtection(
            sheet=True,
            password=self.SHEET_PASSWORD,
            formatCells=False,        # Prevent format changes
            formatColumns=False,      # Prevent column changes
            formatRows=False,         # Prevent row changes  
            insertColumns=False,      # Prevent column insertion
            insertRows=False,         # Prevent row insertion
            insertHyperlinks=False,   # Prevent hyperlinks
            deleteColumns=False,      # Prevent column deletion
            deleteRows=False,         # Prevent row deletion
            selectLockedCells=True,   # Allow selecting locked cells
            selectUnlockedCells=True, # Allow selecting unlocked cells
            sort=False,              # Prevent sorting (would break structure)
            autoFilter=False,        # Prevent filtering (would break structure)
            pivotTables=False        # Prevent pivot tables
        )
        
        # Apply protection
        self.worksheet.protection = protection
        
        # Unlock only the data entry cells (columns A-E, I-K, M for rows 9-200)
        # These are: Xnorm, Ynorm, Znorm, DataID, Cluster, Centroid_X, Centroid_Y, Centroid_Z, Radius
        data_entry_columns = [1, 2, 3, 4, 5, 9, 10, 11, 13]  # A, B, C, D, E, I, J, K, M
        
        for row in range(9, 201):  # Data rows 9-200
            for col in data_entry_columns:
                cell = self.worksheet.cell(row=row, column=col)
                cell.protection = openpyxl.styles.Protection(locked=False)
        
        # Also unlock ∆E column (F) for Plot_3D to write calculated values
        for row in range(9, 201):
            cell = self.worksheet.cell(row=row, column=6)  # Column F (∆E)
            cell.protection = openpyxl.styles.Protection(locked=False)
    
    def _add_instructions(self):
        """Add instructions in a separate sheet."""
        # Create instructions sheet
        instructions_sheet = self.workbook.create_sheet("Instructions")
        
        instructions = [
            ["Plot_3D Template Instructions", ""],
            ["", ""],
            ["CRITICAL FORMAT REQUIREMENTS:", ""],
            ["• Column structure MUST NOT be changed", ""],
            ["• Headers in row 8 are PROTECTED", ""],
            ["• K-means clustering requires exact format", ""],
            ["• ΔE calculations depend on column positions", ""],
            ["", ""],
            ["DATA ENTRY GUIDELINES:", ""],
            ["• Enter data starting at row 9", ""],
            ["• Xnorm, Ynorm, Znorm: Your color coordinates", ""],
            ["• DataID: Unique identifier for each sample", ""],
            ["• Use dropdown menus in gray columns", ""],
            ["• Cluster and ΔE: Calculated by Plot_3D", ""],
            ["", ""],
            ["REFRESH DATA FUNCTIONALITY:", ""],
            ["• Save this file after making changes", ""],
            ["• Use 'Refresh Data' in Plot_3D to reload", ""],
            ["• File format must remain unchanged", ""],
            ["", ""],
            ["PROTECTION INFO:", ""],
            [f"• Sheet is protected with password: {self.SHEET_PASSWORD}", ""],
            ["• Only data entry cells are unlocked", ""],
            ["• Structure cannot be modified", ""],
            ["• This ensures Plot_3D compatibility", ""]
        ]
        
        # Add instructions
        for row_idx, (col1, col2) in enumerate(instructions, 1):
            instructions_sheet.cell(row=row_idx, column=1).value = col1
            instructions_sheet.cell(row=row_idx, column=2).value = col2
            
            # Format headers
            if "CRITICAL" in col1 or "DATA ENTRY" in col1 or "REFRESH" in col1 or "PROTECTION" in col1:
                cell = instructions_sheet.cell(row=row_idx, column=1)
                cell.font = Font(bold=True, color="FF0000")
            elif col1.startswith("Plot_3D"):
                cell = instructions_sheet.cell(row=row_idx, column=1)
                cell.font = Font(bold=True, size=14)
        
        # Set column widths
        instructions_sheet.column_dimensions['A'].width = 40
        instructions_sheet.column_dimensions['B'].width = 20


def create_rigid_plot3d_templates():
    """Create rigid Plot_3D templates in the templates directory."""
    try:
        # Ensure templates directory exists
        templates_dir = "data/templates/plot3d"
        os.makedirs(templates_dir, exist_ok=True)
        
        # Create rigid templates
        template_creator = RigidPlot3DTemplate()
        
        # Create Excel version (full protection)
        xlsx_path = os.path.join(templates_dir, "Plot3D_Rigid_Template.xlsx")
        success_xlsx = template_creator.create_rigid_template(xlsx_path, "Plot3D_Rigid")
        
        # Create ODS version (limited protection but same structure)
        ods_path = os.path.join(templates_dir, "Plot3D_Rigid_Template.ods")
        success_ods = create_ods_rigid_template(ods_path, "Plot3D_Rigid")
        
        results = []
        if success_xlsx:
            results.append(f"✓ Created: {xlsx_path}")
        if success_ods:
            results.append(f"✓ Created: {ods_path}")
            
        return results
        
    except Exception as e:
        logger.error(f"Error creating rigid templates: {e}")
        return [f"✗ Error: {str(e)}"]


def create_ods_rigid_template(file_path: str, sample_set_name: str) -> bool:
    """Create rigid ODS template with same structure (limited protection)."""
    try:
        import pandas as pd
        
        # Create structured data for ODS
        # Metadata rows (1-7)
        metadata = [
            ["Plot_3D Data Template", "", "", "", "", "", "DATA ENTRY AREA:"],
            [f"Sample Set: {sample_set_name}", "", "", "", "", "", "• Enter data starting at row 9"],
            [f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M')}", "", "", "", "", "", "• Use valid values in columns G, H, L"],
            ["", "", "", "", "", "", "• DO NOT modify column headers"],
            ["IMPORTANT: This format is required for Plot_3D", "", "", "", "", "", "• Save file after changes"],
            ["• K-means clustering requires exact structure", "", "", "", "", "", ""],
            ["• ΔE calculations depend on proper format", "", "", "", "", "", ""]
        ]
        
        # Add headers (row 8)
        headers = RigidPlot3DTemplate.PLOT3D_COLUMNS
        
        # Create sample data rows to show structure
        sample_data = []
        for i in range(5):  # Add 5 sample rows
            row = [""] * 13  # 13 columns
            row[3] = f"{sample_set_name}_Sample_{i+1:03d}"  # DataID
            row[6] = "."     # Marker default
            row[7] = "blue"  # Color default
            sample_data.append(row)
        
        # Combine all data
        all_data = metadata + [headers] + sample_data
        
        # Create DataFrame and save as ODS
        df = pd.DataFrame(all_data)
        df.to_excel(file_path, engine='odf', index=False, header=False)
        
        logger.info(f"Created rigid ODS template: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error creating ODS rigid template: {e}")
        return False


if __name__ == "__main__":
    # Test the rigid template creation
    results = create_rigid_plot3d_templates()
    for result in results:
        print(result)
