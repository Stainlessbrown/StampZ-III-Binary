#!/usr/bin/env python3
"""
Data File Manager for StampZ-III

Centralized formatting and rules engine that ensures consistent behavior across:
- Realtime datasheets (tksheet)
- External worksheets (.ods, .xlsx, .csv)
- Plot_3D integration
- Ternary plot integration
- Database import/export

This eliminates the coordination issues by having ONE set of rules that ALL 
formats use consistently.
"""

import logging
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Union, Tuple
from dataclasses import dataclass, field
from enum import Enum
import os
from datetime import datetime

logger = logging.getLogger(__name__)


class DataFormat(Enum):
    """Supported data formats."""
    PLOT3D = "plot3d"
    TERNARY = "ternary"
    REALTIME_SHEET = "realtime_sheet" 
    ODS_EXTERNAL = "ods_external"
    XLSX_EXTERNAL = "xlsx_external"
    CSV_EXTERNAL = "csv_external"
    DATABASE = "database"


@dataclass
class FormatSpecification:
    """Defines the specification for a data format."""
    
    # Column definitions
    columns: List[str]
    required_columns: List[str]
    optional_columns: List[str]
    
    # Validation rules
    validation_lists: Dict[str, List[str]]
    numeric_columns: List[str]
    text_columns: List[str]
    
    # Formatting rules
    protected_areas: List[Tuple[int, int, int, int]]  # (start_row, end_row, start_col, end_col)
    header_row: int
    data_start_row: int
    
    # Colors for UI formatting
    protected_color: str = "#FFB6C1"  # Pink
    validation_color: str = "#E8E8E8"  # Light gray
    marker_color: str = "#FA8072"     # Salmon
    color_color: str = "#FFFF99"      # Yellow
    sphere_color: str = "#FFFF99"     # Yellow


class DataFileManager:
    """
    Centralized manager for all data file formats and formatting rules.
    
    This ensures that Plot_3D, Ternary, realtime sheets, and external files
    all use identical formatting, validation, and data handling rules.
    """
    
    def __init__(self):
        """Initialize the data file manager with format specifications."""
        self._format_specs = self._initialize_format_specifications()
        logger.info("DataFileManager initialized with consistent format specifications")
    
    def _initialize_format_specifications(self) -> Dict[DataFormat, FormatSpecification]:
        """Initialize format specifications for all supported formats."""
        
        # Common validation lists - consistent across ALL formats (empty string = no selection)
        MARKERS = ['', '.', 'o', '*', '^', '<', '>', 'v', 's', 'D', '+', 'x']
        COLORS = [
            '', 'red', 'blue', 'green', 'orange', 'purple', 'yellow',
            'cyan', 'magenta', 'brown', 'pink', 'lime', 'navy', 'teal', 'gray'
        ]
        SPHERES = [
            '', 'red', 'green', 'blue', 'yellow', 'cyan', 'magenta',
            'orange', 'purple', 'brown', 'pink', 'lime', 'navy', 'teal', 'gray'
        ]
        
        specs = {}
        
        # Plot_3D Format Specification
        plot3d_columns = [
            'Xnorm', 'Ynorm', 'Znorm', 'DataID', 'Cluster', 
            'ΔE', 'Marker', 'Color', 'Centroid_X', 'Centroid_Y', 
            'Centroid_Z', 'Sphere', 'Radius', 'Exclude'
        ]
        
        specs[DataFormat.PLOT3D] = FormatSpecification(
            columns=plot3d_columns,
            required_columns=['Xnorm', 'Ynorm', 'Znorm', 'DataID'],
            optional_columns=['Cluster', 'ΔE', 'Marker', 'Color', 'Centroid_X', 'Centroid_Y', 'Centroid_Z', 'Sphere', 'Radius', 'Exclude'],
            validation_lists={
                'Marker': MARKERS,
                'Color': COLORS,
                'Sphere': SPHERES
            },
            numeric_columns=['Xnorm', 'Ynorm', 'Znorm', 'Cluster', 'ΔE', 'Centroid_X', 'Centroid_Y', 'Centroid_Z', 'Radius'],
            text_columns=['DataID', 'Marker', 'Color', 'Sphere', 'Exclude'],
            protected_areas=[],  # No longer reserving rows 2-7 since centroids are dynamically placed starting at row 8
            header_row=0,
            data_start_row=7  # Row 8 in 1-based indexing
        )
        
        # Ternary Format Specification  
        ternary_columns = [
            'L*', 'a*', 'b*', 'DataID', 'Marker', 'Color', 'Group', 'Notes'
        ]
        
        specs[DataFormat.TERNARY] = FormatSpecification(
            columns=ternary_columns,
            required_columns=['L*', 'a*', 'b*', 'DataID'],
            optional_columns=['Marker', 'Color', 'Group', 'Notes'],
            validation_lists={
                'Marker': MARKERS,
                'Color': COLORS
            },
            numeric_columns=['L*', 'a*', 'b*'],
            text_columns=['DataID', 'Marker', 'Color', 'Group', 'Notes'],
            protected_areas=[],  # Ternary has no protected areas
            header_row=0,
            data_start_row=1
        )
        
        # Realtime sheet uses same spec as Plot_3D but with different formatting
        specs[DataFormat.REALTIME_SHEET] = specs[DataFormat.PLOT3D]
        
        # External formats use same column specs as their primary format
        specs[DataFormat.ODS_EXTERNAL] = specs[DataFormat.PLOT3D]
        specs[DataFormat.XLSX_EXTERNAL] = specs[DataFormat.PLOT3D]  
        specs[DataFormat.CSV_EXTERNAL] = specs[DataFormat.PLOT3D]
        
        # Database format uses Plot_3D columns
        specs[DataFormat.DATABASE] = specs[DataFormat.PLOT3D]
        
        return specs
    
    def get_format_spec(self, data_format: DataFormat) -> FormatSpecification:
        """Get the format specification for a given format."""
        return self._format_specs[data_format]
    
    def get_columns(self, data_format: DataFormat) -> List[str]:
        """Get column list for a specific format."""
        return self._format_specs[data_format].columns
    
    def get_validation_lists(self, data_format: DataFormat) -> Dict[str, List[str]]:
        """Get validation lists for a specific format."""
        return self._format_specs[data_format].validation_lists
    
    # === Data Standardization Methods ===
    
    def standardize_data(self, data: Union[pd.DataFrame, List[List[Any]]], 
                        source_format: DataFormat, 
                        target_format: DataFormat) -> Union[pd.DataFrame, List[List[Any]]]:
        """
        Standardize data from one format to another.
        
        Args:
            data: Input data
            source_format: Format of the input data
            target_format: Desired output format
            
        Returns:
            Data in the target format
        """
        try:
            # Convert input to DataFrame for processing
            if isinstance(data, list):
                source_columns = self.get_columns(source_format)
                df = pd.DataFrame(data, columns=source_columns)
            else:
                df = data.copy()
            
            # Get target specification
            target_spec = self.get_format_spec(target_format)
            target_columns = target_spec.columns
            
            # Ensure all target columns exist
            for col in target_columns:
                if col not in df.columns:
                    df[col] = ''
            
            # Reorder columns to match target format
            df = df.reindex(columns=target_columns, fill_value='')
            
            # Apply format-specific transformations
            df = self._apply_format_transformations(df, source_format, target_format)
            
            # Return in requested format
            if isinstance(data, list):
                return df.values.tolist()
            else:
                return df
                
        except Exception as e:
            logger.error(f"Error standardizing data from {source_format} to {target_format}: {e}")
            return data
    
    def _apply_format_transformations(self, df: pd.DataFrame, 
                                    source_format: DataFormat, 
                                    target_format: DataFormat) -> pd.DataFrame:
        """Apply format-specific data transformations."""
        
        # Normalize coordinate data for Plot_3D formats
        if target_format in [DataFormat.PLOT3D, DataFormat.REALTIME_SHEET]:
            df = self._normalize_coordinates_for_plot3d(df)
        
        # Convert coordinate data for Ternary formats
        elif target_format == DataFormat.TERNARY:
            df = self._convert_coordinates_for_ternary(df)
        
        # Clean validation values
        df = self._clean_validation_values(df, target_format)
        
        return df
    
    def _normalize_coordinates_for_plot3d(self, df: pd.DataFrame) -> pd.DataFrame:
        """Normalize coordinate data to 0-1 range for Plot_3D."""
        coord_columns = ['Xnorm', 'Ynorm', 'Znorm']
        
        for col in coord_columns:
            if col in df.columns:
                # Ensure values are in 0-1 range
                df[col] = pd.to_numeric(df[col], errors='coerce')
                df[col] = df[col].clip(0, 1)
        
        return df
    
    def _convert_coordinates_for_ternary(self, df: pd.DataFrame) -> pd.DataFrame:
        """Convert coordinates to L*a*b* format for Ternary."""
        # If we have normalized coordinates, convert back to L*a*b*
        if all(col in df.columns for col in ['Xnorm', 'Ynorm', 'Znorm']):
            df['L*'] = df['Xnorm'] * 100  # 0-1 → 0-100
            df['a*'] = (df['Ynorm'] * 255) - 128  # 0-1 → -128 to +127
            df['b*'] = (df['Znorm'] * 255) - 128  # 0-1 → -128 to +127
        
        return df
    
    def _clean_validation_values(self, df: pd.DataFrame, target_format: DataFormat) -> pd.DataFrame:
        """Clean and validate dropdown values."""
        validation_lists = self.get_validation_lists(target_format)
        
        for column, valid_values in validation_lists.items():
            if column in df.columns:
                # Replace empty/NaN values with empty string
                df[column] = df[column].fillna('')
                df[column] = df[column].replace('(none)', '')
                
                # Validate values are in the allowed list
                mask = ~df[column].isin(valid_values)
                if mask.any():
                    logger.warning(f"Found invalid {column} values, replacing with empty string")
                    df.loc[mask, column] = ''
        
        return df
    
    # === Formatting Application Methods ===
    
    def apply_realtime_sheet_formatting(self, sheet_widget, data_format: DataFormat = DataFormat.PLOT3D):
        """
        Apply consistent formatting to a realtime sheet (tksheet widget).
        
        Args:
            sheet_widget: The tksheet widget to format
            data_format: The data format specification to use
        """
        try:
            print(f"\n🎨 UNIFIED FORMATTING DEBUG: Starting formatting for {data_format}")
            spec = self.get_format_spec(data_format)
            print(f"🎨 FORMAT SPEC: data_start_row={spec.data_start_row}, protected_areas={spec.protected_areas}")
            print(f"🎨 COLORS: marker={spec.marker_color}, color={spec.color_color}, sphere={spec.sphere_color}")
            
            # Get sheet info
            total_rows = sheet_widget.get_total_rows()
            print(f"🎨 SHEET INFO: {total_rows} total rows")
            
            # Clear existing formatting
            try:
                sheet_widget.dehighlight_all()
                print(f"🎨 CLEARED: All existing formatting")
            except Exception as clear_error:
                print(f"🎨 CLEAR ERROR: {clear_error}")
            
            # Apply protected area formatting (pink) - should be empty now
            if spec.protected_areas:
                print(f"🎨 PROTECTED AREAS: Applying {len(spec.protected_areas)} protected areas")
                for start_row, end_row, start_col, end_col in spec.protected_areas:
                    protected_cells = [(row, col) for row in range(start_row, end_row + 1) 
                                     for col in range(start_col, end_col + 1)]
                    sheet_widget.highlight_cells(
                        cells=protected_cells,
                        bg=spec.protected_color,
                        fg='black'
                    )
                    print(f"🎨 PROTECTED: Applied to {len(protected_cells)} cells")
            else:
                print(f"🎨 PROTECTED AREAS: None (as expected)")
            
            # Apply column-specific formatting
            print(f"🎨 COLUMN FORMATTING: Starting...")
            self._apply_column_formatting(sheet_widget, spec)
            
            # Apply validation dropdowns
            print(f"🎨 VALIDATION DROPDOWNS: Starting...")
            self._apply_validation_dropdowns(sheet_widget, spec)
            
            print(f"🎨 UNIFIED FORMATTING: Completed successfully for {data_format}")
            logger.info(f"Applied realtime sheet formatting for {data_format}")
            
        except Exception as e:
            print(f"🎨 UNIFIED FORMATTING ERROR: {e}")
            import traceback
            print(f"🎨 TRACEBACK: {traceback.format_exc()}")
            logger.error(f"Error applying realtime sheet formatting: {e}")
    
    def _apply_column_formatting(self, sheet_widget, spec: FormatSpecification):
        """Apply column-specific formatting to sheet."""
        try:
            # Get current data range - be more conservative to ensure we catch all data
            total_rows = sheet_widget.get_total_rows()
            # Find the last row with actual data to avoid formatting empty rows
            last_data_row = total_rows
            for row in range(total_rows - 1, spec.data_start_row - 1, -1):  # Start from end, stop at data start
                try:
                    # Check if any cell in this row has data
                    row_data = [sheet_widget.get_cell_data(row, col) for col in range(len(spec.columns))]
                    if any(cell and str(cell).strip() for cell in row_data):
                        last_data_row = row + 1
                        break
                except Exception:
                    continue
            
            columns = spec.columns
            
            # Format Marker column (salmon) - only for data rows
            if 'Marker' in columns and last_data_row > spec.data_start_row:
                marker_col = columns.index('Marker')
                marker_cells = [(row, marker_col) for row in range(spec.data_start_row, last_data_row)]
                sheet_widget.highlight_cells(cells=marker_cells, bg=spec.marker_color, fg='black')
                logger.info(f"Applied marker column formatting: rows {spec.data_start_row}-{last_data_row-1}")
            
            # Format Color column (yellow) - only for data rows  
            if 'Color' in columns and last_data_row > spec.data_start_row:
                color_col = columns.index('Color')
                color_cells = [(row, color_col) for row in range(spec.data_start_row, last_data_row)]
                sheet_widget.highlight_cells(cells=color_cells, bg=spec.color_color, fg='black')
                logger.info(f"Applied color column formatting: rows {spec.data_start_row}-{last_data_row-1}")
            
            # Format Sphere column (yellow) - start from row 2 since sphere data can be in centroid area
            if 'Sphere' in columns and last_data_row > 1:
                sphere_col = columns.index('Sphere')
                # Sphere formatting starts from row 2 (index 1) since centroids can have spheres
                sphere_start = 1  # Always start from row 2 (index 1)
                sphere_cells = [(row, sphere_col) for row in range(sphere_start, last_data_row)]
                sheet_widget.highlight_cells(cells=sphere_cells, bg=spec.sphere_color, fg='black')
                logger.info(f"Applied sphere column formatting: rows {sphere_start+1}-{last_data_row}")
            
            # Apply center alignment to all data cells
            if last_data_row > 0:
                all_cells = [(row, col) for row in range(last_data_row) for col in range(len(columns))]
                sheet_widget.align_cells(cells=all_cells, align='center')
                logger.info(f"Applied center alignment to {len(all_cells)} cells")
            
        except Exception as e:
            logger.error(f"Error applying column formatting: {e}")
            import traceback
            logger.error(f"Column formatting traceback: {traceback.format_exc()}")
    
    def _apply_validation_dropdowns(self, sheet_widget, spec: FormatSpecification):
        """Apply validation dropdowns to sheet."""
        try:
            total_rows = sheet_widget.get_total_rows()
            columns = spec.columns
            validation_lists = spec.validation_lists
            
            print(f"DEBUG: Setting up validation dropdowns for {len(validation_lists)} columns, {total_rows} rows")
            
            for column_name, valid_values in validation_lists.items():
                if column_name in columns:
                    col_index = columns.index(column_name)
                    
                    # Determine row range for this column
                    if column_name == 'Sphere':
                        # Sphere dropdowns: start from row 2 (index 1) for centroid area
                        start_row = 1  # Row 2 in display
                        end_row = total_rows
                        print(f"DEBUG: {column_name} dropdowns: rows {start_row+1}-{end_row} (centroid + data area)")
                    elif column_name in ['Marker', 'Color']:
                        # Marker/Color dropdowns: start from data area only (row 8)
                        start_row = spec.data_start_row  # Row 8 in display
                        end_row = total_rows
                        print(f"DEBUG: {column_name} dropdowns: rows {start_row+1}-{end_row} (data area only)")
                    else:
                        # Other dropdowns: start from data area
                        start_row = spec.data_start_row
                        end_row = total_rows
                        print(f"DEBUG: {column_name} dropdowns: rows {start_row+1}-{end_row} (data area)")
                    
                    dropdown_count = 0
                    
                    # Create dropdowns for this column
                    for row in range(start_row, end_row):
                        try:
                            # Get current value or default
                            current_value = sheet_widget.get_cell_data(row, col_index)
                            if not current_value or current_value.strip() == '':
                                # Use smart default based on coordinate data and column type
                                if self._has_coordinate_data(sheet_widget, row):
                                    # Has coordinates - use appropriate default
                                    if column_name == 'Marker':
                                        current_value = '.'  # Default marker
                                    elif column_name == 'Color':
                                        current_value = 'blue'  # Default color
                                    elif column_name == 'Sphere':
                                        current_value = ''  # Empty sphere by default
                                    else:
                                        current_value = valid_values[1] if len(valid_values) > 1 else valid_values[0]
                                else:
                                    # No coordinates - empty default
                                    current_value = ''
                            
                            # Validate current value is in the allowed list
                            if current_value and current_value not in valid_values:
                                print(f"DEBUG: Invalid {column_name} '{current_value}' at row {row}, using empty")
                                current_value = ''
                            
                            sheet_widget.create_dropdown(
                                r=row, c=col_index,
                                values=valid_values,
                                set_value=current_value,
                                redraw=False
                            )
                            dropdown_count += 1
                            
                        except Exception as dropdown_error:
                            logger.debug(f"Error creating dropdown at row {row}, col {col_index}: {dropdown_error}")
                    
                    print(f"DEBUG: Created {dropdown_count} {column_name} dropdowns")
            
            # Refresh sheet to show dropdowns
            sheet_widget.refresh()
            print("DEBUG: Sheet refreshed to show dropdowns")
            
        except Exception as e:
            logger.error(f"Error applying validation dropdowns: {e}")
            import traceback
            logger.error(f"Validation dropdown traceback: {traceback.format_exc()}")
    
    def _has_coordinate_data(self, sheet_widget, row: int) -> bool:
        """Check if a row has coordinate data (Xnorm, Ynorm, Znorm)."""
        try:
            # Check first 3 columns (Xnorm, Ynorm, Znorm)
            for col in range(3):
                cell_data = sheet_widget.get_cell_data(row, col)
                if cell_data and str(cell_data).strip():
                    return True
            return False
        except Exception:
            return False
    
    # === External File Methods ===
    
    def create_external_file(self, file_path: str, data: Optional[pd.DataFrame] = None, 
                           data_format: DataFormat = DataFormat.PLOT3D) -> bool:
        """
        Create an external file with consistent formatting.
        
        Args:
            file_path: Path to create the file
            data: Optional data to populate the file
            data_format: Format specification to use
            
        Returns:
            True if successful
        """
        try:
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.ods':
                return self._create_ods_file(file_path, data, data_format)
            elif file_ext in ['.xlsx', '.xls']:
                return self._create_xlsx_file(file_path, data, data_format)
            elif file_ext == '.csv':
                return self._create_csv_file(file_path, data, data_format)
            else:
                logger.error(f"Unsupported file format: {file_ext}")
                return False
                
        except Exception as e:
            logger.error(f"Error creating external file: {e}")
            return False
    
    def _create_ods_file(self, file_path: str, data: Optional[pd.DataFrame], 
                        data_format: DataFormat) -> bool:
        """Create ODS file with consistent formatting."""
        try:
            from odf.opendocument import OpenDocumentSpreadsheet
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
            
            spec = self.get_format_spec(data_format)
            doc = OpenDocumentSpreadsheet()
            table = Table(name="StampZ_Data")
            
            # Add header row
            header_row = TableRow()
            for col_name in spec.columns:
                cell = TableCell()
                cell.addElement(P(text=col_name))
                header_row.addElement(cell)
            table.addElement(header_row)
            
            # Add data rows if provided
            if data is not None:
                standardized_data = self.standardize_data(data, data_format, data_format)
                for _, row in standardized_data.iterrows():
                    data_row = TableRow()
                    for col_name in spec.columns:
                        cell = TableCell()
                        value = row[col_name] if col_name in row else ''
                        cell.addElement(P(text=str(value)))
                        data_row.addElement(cell)
                    table.addElement(data_row)
            
            doc.spreadsheet.addElement(table)
            doc.save(file_path)
            
            logger.info(f"Created ODS file: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating ODS file: {e}")
            return False
    
    def _create_xlsx_file(self, file_path: str, data: Optional[pd.DataFrame], 
                         data_format: DataFormat) -> bool:
        """Create XLSX file with consistent formatting."""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
            
            spec = self.get_format_spec(data_format)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "StampZ_Data"
            
            # Add headers
            for col_idx, col_name in enumerate(spec.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Add data if provided
            if data is not None:
                standardized_data = self.standardize_data(data, data_format, data_format)
                for row_idx, (_, row) in enumerate(standardized_data.iterrows(), 2):
                    for col_idx, col_name in enumerate(spec.columns, 1):
                        value = row[col_name] if col_name in row else ''
                        ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply formatting
            self._apply_xlsx_formatting(ws, spec)
            
            wb.save(file_path)
            logger.info(f"Created XLSX file: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating XLSX file: {e}")
            return False
    
    def _apply_xlsx_formatting(self, worksheet, spec: FormatSpecification):
        """Apply consistent formatting to XLSX worksheet."""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
            
            # Apply protected area formatting
            protected_fill = PatternFill(start_color=spec.protected_color[1:], end_color=spec.protected_color[1:], fill_type='solid')
            
            for start_row, end_row, start_col, end_col in spec.protected_areas:
                for row in range(start_row + 1, end_row + 2):  # Convert to 1-based
                    for col in range(start_col + 1, end_col + 2):
                        worksheet.cell(row=row, column=col).fill = protected_fill
            
            # Apply validation formatting
            validation_fill = PatternFill(start_color=spec.validation_color[1:], end_color=spec.validation_color[1:], fill_type='solid')
            
            columns = spec.columns
            for column_name in spec.validation_lists.keys():
                if column_name in columns:
                    col_idx = columns.index(column_name) + 1  # 1-based
                    start_row = spec.data_start_row + 1 if column_name != 'Sphere' else 2
                    
                    # Apply to a reasonable range (e.g., 100 rows)
                    for row in range(start_row, start_row + 100):
                        worksheet.cell(row=row, column=col_idx).fill = validation_fill
            
        except Exception as e:
            logger.error(f"Error applying XLSX formatting: {e}")
    
    def _create_csv_file(self, file_path: str, data: Optional[pd.DataFrame], 
                        data_format: DataFormat) -> bool:
        """Create CSV file with consistent structure."""
        try:
            spec = self.get_format_spec(data_format)
            
            if data is not None:
                standardized_data = self.standardize_data(data, data_format, data_format)
            else:
                standardized_data = pd.DataFrame(columns=spec.columns)
            
            standardized_data.to_csv(file_path, index=False)
            logger.info(f"Created CSV file: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating CSV file: {e}")
            return False
    
    def read_external_file(self, file_path: str, data_format: DataFormat) -> Optional[pd.DataFrame]:
        """Read external file and return standardized data."""
        try:
            file_ext = os.path.splitext(file_path)[1].lower()
            
            if file_ext == '.ods':
                data = self._read_ods_file(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                data = self._read_xlsx_file(file_path)
            elif file_ext == '.csv':
                data = self._read_csv_file(file_path)
            else:
                logger.error(f"Unsupported file format: {file_ext}")
                return None
            
            if data is not None:
                # Standardize the data to the requested format
                return self.standardize_data(data, data_format, data_format)
            
            return None
            
        except Exception as e:
            logger.error(f"Error reading external file: {e}")
            return None
    
    def _read_ods_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """Read ODS file."""
        try:
            # Try pandas first (requires odfpy)
            return pd.read_excel(file_path, engine='odf')
        except Exception as e:
            logger.error(f"Error reading ODS file: {e}")
            return None
    
    def _read_xlsx_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """Read XLSX file."""
        try:
            return pd.read_excel(file_path, engine='openpyxl')
        except Exception as e:
            logger.error(f"Error reading XLSX file: {e}")
            return None
    
    def _read_csv_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """Read CSV file."""
        try:
            return pd.read_csv(file_path)
        except Exception as e:
            logger.error(f"Error reading CSV file: {e}")
            return None
    
    # === Utility Methods ===
    
    def validate_data(self, data: pd.DataFrame, data_format: DataFormat) -> Dict[str, List[str]]:
        """
        Validate data against format specification.
        
        Returns:
            Dictionary of column_name -> list of validation errors
        """
        errors = {}
        spec = self.get_format_spec(data_format)
        
        # Check required columns
        missing_required = set(spec.required_columns) - set(data.columns)
        if missing_required:
            errors['missing_columns'] = [f"Missing required columns: {list(missing_required)}"]
        
        # Validate column data types
        for col in spec.numeric_columns:
            if col in data.columns:
                non_numeric = data[col].apply(lambda x: not pd.isna(x) and not str(x).strip() in ['', '(none)'] and not pd.api.types.is_numeric_dtype(type(x)))
                if non_numeric.any():
                    errors[col] = errors.get(col, []) + [f"Non-numeric values found in numeric column"]
        
        # Validate dropdown values
        for col, valid_values in spec.validation_lists.items():
            if col in data.columns:
                invalid_values = ~data[col].isin(valid_values + ['', np.nan])
                if invalid_values.any():
                    errors[col] = errors.get(col, []) + [f"Invalid values found: {list(data[col][invalid_values].unique())}"]
        
        return errors
    
    def get_format_info(self, data_format: DataFormat) -> Dict[str, Any]:
        """Get comprehensive information about a format."""
        spec = self.get_format_spec(data_format)
        
        return {
            'format': data_format.value,
            'columns': spec.columns,
            'required_columns': spec.required_columns,
            'validation_lists': spec.validation_lists,
            'header_row': spec.header_row,
            'data_start_row': spec.data_start_row,
            'protected_areas': spec.protected_areas
        }


# === Global Manager Instance ===

_global_manager = None

def get_data_file_manager() -> DataFileManager:
    """Get the global data file manager instance."""
    global _global_manager
    if _global_manager is None:
        _global_manager = DataFileManager()
    return _global_manager


# === Convenience Functions ===

def format_data_for_plot3d(data: Union[pd.DataFrame, List[List[Any]]]) -> pd.DataFrame:
    """Format data consistently for Plot_3D use."""
    manager = get_data_file_manager()
    return manager.standardize_data(data, DataFormat.DATABASE, DataFormat.PLOT3D)

def format_data_for_ternary(data: Union[pd.DataFrame, List[List[Any]]]) -> pd.DataFrame:
    """Format data consistently for Ternary use."""
    manager = get_data_file_manager()
    return manager.standardize_data(data, DataFormat.DATABASE, DataFormat.TERNARY)

def apply_realtime_formatting(sheet_widget, format_type: str = 'plot3d'):
    """Apply consistent formatting to realtime sheet."""
    manager = get_data_file_manager()
    data_format = DataFormat.PLOT3D if format_type == 'plot3d' else DataFormat.TERNARY
    manager.apply_realtime_sheet_formatting(sheet_widget, data_format)

def create_external_worksheet(file_path: str, data: Optional[pd.DataFrame] = None, 
                             format_type: str = 'plot3d') -> bool:
    """Create external worksheet with consistent formatting."""
    manager = get_data_file_manager()
    data_format = DataFormat.PLOT3D if format_type == 'plot3d' else DataFormat.TERNARY
    return manager.create_external_file(file_path, data, data_format)


if __name__ == "__main__":
    # Test the data file manager
    manager = get_data_file_manager()
    
    # Test format specifications
    plot3d_spec = manager.get_format_spec(DataFormat.PLOT3D)
    print(f"Plot_3D columns: {plot3d_spec.columns}")
    print(f"Plot_3D validation: {plot3d_spec.validation_lists}")
    
    # Test data standardization
    test_data = pd.DataFrame({
        'Xnorm': [0.5, 0.6],
        'Ynorm': [0.7, 0.8], 
        'Znorm': [0.9, 0.4],
        'DataID': ['Test1', 'Test2'],
        'Marker': ['o', '*'],
        'Color': ['red', 'blue']
    })
    
    standardized = manager.standardize_data(test_data, DataFormat.DATABASE, DataFormat.PLOT3D)
    print(f"Standardized data shape: {standardized.shape}")
    print(f"Standardized columns: {list(standardized.columns)}")