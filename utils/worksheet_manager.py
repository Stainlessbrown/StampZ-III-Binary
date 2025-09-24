"""
Worksheet Manager for StampZ-III

Creates properly formatted Excel worksheets with data validation and 
visual formatting for Plot_3D integration.
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
import logging
from utils.rigid_plot3d_templates import RigidPlot3DTemplate

# Optional ODS support via odfpy for precise structure control
try:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    ODF_AVAILABLE = True
except ImportError:
    ODF_AVAILABLE = False

logger = logging.getLogger(__name__)


class WorksheetManager:
    """Manages Excel worksheet creation with Plot_3D formatting and validation."""
    
    # Plot_3D column structure
    PLOT3D_COLUMNS = [
        'Xnorm', 'Ynorm', 'Znorm', 'DataID', 'Cluster', 
        '∆E', 'Marker', 'Color', 'Centroid_X', 'Centroid_Y', 
        'Centroid_Z', 'Sphere', 'Radius'
    ]
    
    # Data validation lists from Plot_3D code
    VALID_MARKERS = ['.', 'o', '*', '^', '<', '>', 'v', 's', 'D', '+', 'x']
    
    VALID_COLORS = [
        'red', 'blue', 'green', 'orange', 'purple', 'yellow', 
        'cyan', 'magenta', 'brown', 'pink', 'lime', 'navy', 'teal', 'gray'
    ]
    
    VALID_SPHERES = [
        'red', 'green', 'blue', 'yellow', 'cyan', 'magenta', 
        'orange', 'purple', 'brown', 'pink', 'lime', 'navy', 'teal', 'gray'
    ]
    
    def __init__(self):
        """Initialize the worksheet manager."""
        self.workbook = None
        self.worksheet = None
        self.rigid_template_creator = RigidPlot3DTemplate()
        
    def create_plot3d_worksheet(self, file_path: str, sample_set_name: str = "StampZ_Analysis") -> bool:
        """
        Create a new Excel worksheet formatted for Plot_3D with data validation.
        
        Args:
            file_path: Path where to save the Excel file
            sample_set_name: Name for the sample set (used in DataID generation)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Create new workbook and worksheet
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Plot3D_Data"
            
            # Set up column headers (row 1)
            self._setup_headers()
            
            # Apply formatting and protection
            self._apply_visual_formatting()
            
            # Set up data validation
            self._setup_data_validation()
            
            # Save the workbook
            self.workbook.save(file_path)
            logger.info(f"Created Plot_3D worksheet: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating worksheet: {e}")
            return False
    
    def _create_simple_plot3d_template(self, file_path: str, sample_set_name: str = "StampZ_Analysis") -> bool:
        """
        Create a Plot_3D ODS template that follows rigid format rules:
        - Rows 1-7: metadata/instructions
        - Row 8: Plot_3D column headers (PLOT3D_COLUMNS)
        - Row 9+: Data rows
        """
        try:
            if not ODF_AVAILABLE:
                raise ImportError("odfpy not available. Install with: pip install odfpy==1.4.1")
            
            doc = OpenDocumentSpreadsheet()
            table = Table(name="Plot3D_Data")
            
            # Rows 1-7 metadata/instructions
            meta_rows = [
                ["Plot_3D Data Template", "", "", "", "", "", "", "", "", "", "", "", "Instructions"],
                [f"Sample Set: {sample_set_name}", "", "", "", "", "", "", "", "", "", "", "", "Enter data starting at row 9"],
                ["Created by StampZ-III", "", "", "", "", "", "", "", "", "", "", "", "Use dropdowns where applicable"],
                ["", "", "", "", "", "", "", "", "", "", "", "", "Do NOT modify headers"],
                ["IMPORTANT: This format is required for Plot_3D", "", "", "", "", "", "", "", "", "", "", "", "Save before Refresh Data"],
                ["K-means expects exact column order", "", "", "", "", "", "", "", "", "", "", "", ""],
                ["ΔE calculations depend on structure", "", "", "", "", "", "", "", "", "", "", "", ""],
            ]
            for row_vals in meta_rows:
                tr = TableRow()
                for val in row_vals:
                    tc = TableCell()
                    tc.addElement(P(text=str(val)))
                    tr.addElement(tc)
                table.addElement(tr)
            
            # Row 8 headers
            header_tr = TableRow()
            for header in self.PLOT3D_COLUMNS:
                tc = TableCell()
                tc.addElement(P(text=header))
                header_tr.addElement(tc)
            table.addElement(header_tr)
            
            # Add 3 example rows (optional)
            for i in range(3):
                tr = TableRow()
                example = {
                    'Xnorm': '', 'Ynorm': '', 'Znorm': '',
                    'DataID': f"{sample_set_name}_Sample_{i+1:03d}",
                    'Cluster': '', '∆E': '',
                    'Marker': '.', 'Color': 'blue',
                    'Centroid_X': '', 'Centroid_Y': '', 'Centroid_Z': '',
                    'Sphere': '', 'Radius': ''
                }
                for col in self.PLOT3D_COLUMNS:
                    tc = TableCell()
                    tc.addElement(P(text=str(example.get(col, ''))))
                    tr.addElement(tc)
                table.addElement(tr)
            
            doc.spreadsheet.addElement(table)
            doc.save(file_path)
            logger.info(f"Created Plot_3D ODS template with rigid layout: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Error creating ODS rigid template: {e}")
            return False
    
    def _setup_headers(self):
        """Set up column headers in row 1."""
        for col_idx, header in enumerate(self.PLOT3D_COLUMNS, 1):
            cell = self.worksheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    
    def _apply_visual_formatting(self):
        """Apply visual formatting including protected areas."""
        # Pink fill for protected data entry area (A2:H7)
        # This covers Xnorm through Color columns, rows 2-7
        pink_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
        
        for row in range(2, 8):  # Rows 2-7
            for col in range(1, 9):  # Columns A-H (1-8)
                cell = self.worksheet.cell(row=row, column=col)
                cell.fill = pink_fill
        
        # Light gray fill for data validation areas
        # G8:H107 (Marker and Color validation)
        # L2:L107 (Sphere validation) 
        gray_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
        
        # Marker column (G) - rows 8-107
        for row in range(8, 108):
            cell = self.worksheet.cell(row=row, column=7)  # Column G
            cell.fill = gray_fill
            
        # Color column (H) - rows 8-107  
        for row in range(8, 108):
            cell = self.worksheet.cell(row=row, column=8)  # Column H
            cell.fill = gray_fill
            
        # Sphere column (L) - rows 2-107
        for row in range(2, 108):
            cell = self.worksheet.cell(row=row, column=12)  # Column L
            cell.fill = gray_fill
    
    def _setup_data_validation(self):
        """Set up data validation dropdown lists."""
        # Marker validation (Column G, rows 8-107)
        marker_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(self.VALID_MARKERS)}"',
            allow_blank=True,
            showDropDown=True
        )
        marker_validation.error = "Please select a valid marker"
        marker_validation.errorTitle = "Invalid Marker"
        self.worksheet.add_data_validation(marker_validation)
        marker_validation.add(f"G8:G107")
        
        # Color validation (Column H, rows 8-107)
        color_validation = DataValidation(
            type="list", 
            formula1=f'"{",".join(self.VALID_COLORS)}"',
            allow_blank=True,
            showDropDown=True
        )
        color_validation.error = "Please select a valid color"
        color_validation.errorTitle = "Invalid Color"
        self.worksheet.add_data_validation(color_validation)
        color_validation.add(f"H8:H107")
        
        # Sphere validation (Column L, rows 2-107)
        sphere_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(self.VALID_SPHERES)}"',
            allow_blank=True,
            showDropDown=True
        )
        sphere_validation.error = "Please select a valid sphere color"
        sphere_validation.errorTitle = "Invalid Sphere Color"
        self.worksheet.add_data_validation(sphere_validation)
        sphere_validation.add(f"L2:L107")
    
    def load_stampz_data(self, sample_set_name: str) -> bool:
        """
        Load data from StampZ database and populate the worksheet.
        
        Args:
            sample_set_name: Name of the sample set to load
            
        Returns:
            True if successful, False otherwise
        """
        try:
            from utils.color_analysis_db import ColorAnalysisDB
            from utils.user_preferences import UserPreferences
            
            # Get measurements from database
            db = ColorAnalysisDB(sample_set_name)
            measurements = db.get_all_measurements()
            
            if not measurements:
                logger.warning(f"No measurements found for sample set: {sample_set_name}")
                return False
            
            # Check user normalization preference
            prefs = UserPreferences()
            export_normalized = prefs.get_export_normalized_values()
            
            # Convert measurements to worksheet format
            self._populate_from_measurements(measurements, sample_set_name, export_normalized)
            
            logger.info(f"Loaded {len(measurements)} measurements into worksheet")
            return True
            
        except Exception as e:
            logger.error(f"Error loading StampZ data: {e}")
            return False
    
    def _populate_from_measurements(self, measurements: List[Dict], sample_set_name: str, export_normalized: bool):
        """Populate worksheet with measurement data."""
        # Start data at row 8 (after protected header area)
        start_row = 8
        
        for i, measurement in enumerate(measurements):
            row = start_row + i
            
            # Get Lab values - these are raw L*a*b* values from database that need normalization
            l_val = measurement.get('l_value', 0.0)
            a_val = measurement.get('a_value', 0.0)  
            b_val = measurement.get('b_value', 0.0)
            
            # CRITICAL FIX: Database stores raw L*a*b* values, but Plot_3D requires 0-1 normalized values
            # Apply proper normalization to convert from raw color space to Plot_3D format
            # L*: 0-100 → 0-1
            # a*: -128 to +127 → 0-1 
            # b*: -128 to +127 → 0-1
            
            # Normalize values regardless of export_normalized preference for Plot_3D compatibility
            x_norm = max(0.0, min(1.0, (l_val if l_val is not None else 0.0) / 100.0))
            y_norm = max(0.0, min(1.0, ((a_val if a_val is not None else 0.0) + 128.0) / 255.0))
            z_norm = max(0.0, min(1.0, ((b_val if b_val is not None else 0.0) + 128.0) / 255.0))
            
            self.worksheet.cell(row=row, column=1).value = round(x_norm, 4)  # Xnorm (normalized L*)
            self.worksheet.cell(row=row, column=2).value = round(y_norm, 4)  # Ynorm (normalized a*)  
            self.worksheet.cell(row=row, column=3).value = round(z_norm, 4)  # Znorm (normalized b*)
            self.worksheet.cell(row=row, column=4).value = f"{sample_set_name}_Sample_{i+1:03d}"  # DataID
            
            # Default marker and color values
            self.worksheet.cell(row=row, column=7).value = '.'  # Marker (default dot)
            self.worksheet.cell(row=row, column=8).value = 'blue'  # Color (default blue)
    
    def save_worksheet(self, file_path: str) -> bool:
        """
        Save the current worksheet to file.
        
        Args:
            file_path: Path where to save the Excel file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if self.workbook:
                self.workbook.save(file_path)
                logger.info(f"Saved worksheet to: {file_path}")
                return True
            else:
                logger.error("No workbook to save")
                return False
                
        except Exception as e:
            logger.error(f"Error saving worksheet: {e}")
            return False
    
    def export_to_format(self, file_path: str, export_format: str) -> bool:
        """
        Export worksheet data to different formats (.ods, .xlsx, .csv).
        
        Args:
            file_path: Path where to save the file
            export_format: Format to export ('ods', 'xlsx', 'csv')
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not self.worksheet:
                logger.error("No worksheet data to export")
                return False
            
            # Convert worksheet to pandas DataFrame for flexible export
            import pandas as pd
            
            # Extract data from worksheet
            data = []
            for row in self.worksheet.iter_rows(min_row=1, values_only=True):
                if any(cell is not None for cell in row):  # Skip completely empty rows
                    data.append(row)
            
            if not data:
                logger.warning("No data to export")
                return False
            
            # Create DataFrame
            df = pd.DataFrame(data[1:], columns=data[0])  # First row is headers
            
            # Export based on format
            if export_format.lower() == 'csv':
                df.to_csv(file_path, index=False)
            elif export_format.lower() == 'ods':
                df.to_excel(file_path, engine='odf', index=False)
            elif export_format.lower() == 'xlsx':
                # For xlsx, save the formatted workbook to preserve formatting
                self.workbook.save(file_path)
            else:
                raise ValueError(f"Unsupported export format: {export_format}")
            
            logger.info(f"Exported worksheet to {export_format.upper()}: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to {export_format}: {e}")
            return False
    
    def create_rigid_plot3d_worksheet(self, file_path: str, sample_set_name: str = "StampZ_Analysis") -> bool:
        """
        Create a rigid, protected Plot_3D worksheet that prevents format corruption.
        
        This method creates templates with:
        - Protected column structure required for K-means clustering
        - Format compliance for ΔE calculations
        - "Refresh Data" functionality support
        - Data validation dropdowns
        - Sheet protection preventing structural changes
        
        Args:
            file_path: Path where to save the Excel file
            sample_set_name: Name for the sample set
            
        Returns:
            True if successful, False otherwise
        """
        try:
            success = self.rigid_template_creator.create_rigid_template(file_path, sample_set_name)
            if success:
                # Set up this manager to work with the created template
                self.workbook = openpyxl.load_workbook(file_path)
                self.worksheet = self.workbook['Plot3D_Data']
                logger.info(f"Created rigid Plot_3D worksheet: {file_path}")
            return success
            
        except Exception as e:
            logger.error(f"Error creating rigid worksheet: {e}")
            return False
    
    def populate_rigid_template(self, sample_set_name: str) -> bool:
        """
        Populate a rigid template with StampZ data while maintaining protection.
        
        Args:
            sample_set_name: Name of the sample set to load data from
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not self.worksheet:
                logger.error("No worksheet loaded. Create or load a rigid template first.")
                return False
                
            # Temporarily unprotect for data population (if needed)
            sheet_was_protected = self.worksheet.protection.sheet
            if sheet_was_protected:
                self.worksheet.protection.sheet = False
            
            # Load and populate data
            success = self.load_stampz_data(sample_set_name)
            
            # Restore protection
            if sheet_was_protected:
                self.worksheet.protection.sheet = True
                
            return success
            
        except Exception as e:
            logger.error(f"Error populating rigid template: {e}")
            return False
    
    def is_rigid_template(self, file_path: str) -> bool:
        """
        Check if a file is a rigid Plot_3D template.
        
        Args:
            file_path: Path to the file to check
            
        Returns:
            True if it's a rigid template, False otherwise
        """
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            # Check for rigid template markers
            cell_a1 = worksheet['A1'].value
            cell_a5 = worksheet['A5'].value
            
            is_rigid = (
                cell_a1 and "Plot_3D Data Template" in str(cell_a1) and
                cell_a5 and "IMPORTANT: This template format is required" in str(cell_a5)
            )
            
            workbook.close()
            return is_rigid
            
        except Exception as e:
            logger.error(f"Error checking if file is rigid template: {e}")
            return False


# Example usage
def create_plot3d_template(file_path: str, sample_set_name: str = "StampZ_Analysis") -> bool:
    """
    Create a Plot_3D template with proper formatting and validation.
    
    Args:
        file_path: Where to save the Excel template
        sample_set_name: Name for the sample set
        
    Returns:
        True if successful, False otherwise
    """
    manager = WorksheetManager()
    success = manager.create_plot3d_worksheet(file_path, sample_set_name)
    
    if success:
        # Optionally load data from StampZ
        # manager.load_stampz_data(sample_set_name)
        # manager.save_worksheet(file_path)
        pass
    
    return success


if __name__ == "__main__":
    # Test the worksheet manager
    test_file = "/tmp/plot3d_template.xlsx"
    if create_plot3d_template(test_file):
        print(f"Template created successfully: {test_file}")
    else:
        print("Failed to create template")
