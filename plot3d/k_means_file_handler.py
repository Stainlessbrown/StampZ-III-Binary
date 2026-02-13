"""
K-means File Handler

Handles saving cluster assignments to both ODS (LibreOffice) and XLSX (Excel) files.
Includes file locking, backup creation, verification, and detailed error handling.
"""

import os
import shutil
import fcntl
import errno
import time
import pandas as pd
from tkinter import messagebox
import logging


class KMeansFileHandler:
    """Handles saving K-means cluster assignments to spreadsheet files."""
    
    def __init__(self, data, file_path, logger=None, sheet_name=None):
        """
        Initialize the file handler.
        
        Args:
            data: pandas DataFrame containing cluster assignments
            file_path: Path to the spreadsheet file (.ods or .xlsx)
            logger: Logger instance for debugging
            sheet_name: Name of the sheet to write to (for multi-sheet files)
        """
        self.data = data
        self.file_path = file_path
        self.logger = logger or logging.getLogger(__name__)
        self.sheet_name = sheet_name
    
    def save_clusters(self, start: int, end: int, row_indices) -> bool:
        """
        Save cluster assignments to the appropriate file format.
        
        Args:
            start: Start row for clustering
            end: End row for clustering
            row_indices: List of row indices to update
            
        Returns:
            True if save was successful
        """
        try:
            if self.file_path.endswith('.xlsx'):
                self.logger.info(f"Saving clusters to Excel file: {self.file_path}")
                return self._save_to_xlsx(start, end, row_indices)
            elif self.file_path.endswith('.ods'):
                self.logger.info(f"Saving clusters to ODS file: {self.file_path}")
                return self._save_to_ods(start, end, row_indices)
            else:
                raise ValueError(f"Unsupported file format: {self.file_path}")
        except Exception as e:
            self.logger.error(f"Error saving clusters: {str(e)}")
            raise
    
    def _save_to_xlsx(self, start: int, end: int, row_indices) -> bool:
        """Save cluster assignments to an Excel (.xlsx) file."""
        try:
            from openpyxl import load_workbook
            
            # Inform user of the process
            msg = (f"Saving cluster assignments for rows {start}-{end}:\\n\\n"
                  "1. Your original .xlsx file will be updated directly\\n"
                  "2. Cluster and Centroid columns will be modified\\n\\n"
                  "Continue?")
            
            if not messagebox.askokcancel("Save Clusters", msg):
                return False
            
            # Create backup
            backup_path = f"{self.file_path}.bak"
            self.logger.info(f"Creating backup at: {backup_path}")
            shutil.copy2(self.file_path, backup_path)
            
            try:
                # Load the Excel file
                wb = load_workbook(self.file_path)
                # Use specified sheet or default to active
                if self.sheet_name and self.sheet_name in wb.sheetnames:
                    ws = wb[self.sheet_name]
                    self.logger.info(f"Using specified sheet: {self.sheet_name}")
                else:
                    ws = wb.active
                    if self.sheet_name:
                        self.logger.warning(f"Sheet '{self.sheet_name}' not found, using active sheet: {ws.title}")
                self.logger.info(f"Opened Excel sheet: {ws.title}")
                
                # Get cluster assignments
                clusters = self.data.iloc[row_indices]['Cluster']
                valid_clusters = clusters[clusters.notna()]
                
                if not valid_clusters.empty:
                    # Find column indices
                    cluster_col = None
                    centroid_x_col = None
                    centroid_y_col = None
                    centroid_z_col = None
                    
                    for col_idx, cell in enumerate(ws[1], 1):
                        if cell.value == 'Cluster':
                            cluster_col = col_idx
                        elif cell.value == 'Centroid_X':
                            centroid_x_col = col_idx
                        elif cell.value == 'Centroid_Y':
                            centroid_y_col = col_idx
                        elif cell.value == 'Centroid_Z':
                            centroid_z_col = col_idx
                    
                    if not all([cluster_col, centroid_x_col, centroid_y_col, centroid_z_col]):
                        missing = []
                        if not cluster_col: missing.append('Cluster')
                        if not centroid_x_col: missing.append('Centroid_X')
                        if not centroid_y_col: missing.append('Centroid_Y')
                        if not centroid_z_col: missing.append('Centroid_Z')
                        raise ValueError(f"Required columns not found: {', '.join(missing)}")
                    
                    self.logger.info(f"Found columns - Cluster: {cluster_col}, Centroid_X: {centroid_x_col}, Centroid_Y: {centroid_y_col}, Centroid_Z: {centroid_z_col}")
                    
                    # Calculate centroids
                    cluster_centroids = self._calculate_centroids()
                    
                    # Write cluster assignments to data rows
                    # Note: row_indices are 0-based DataFrame indices
                    # Sheet rows: row 1 = header, row 2+ = data (DataFrame indices 0+)
                    cluster_write_count = 0
                    for i, idx in enumerate(row_indices):
                        cluster_value = clusters.iloc[i]
                        if pd.notna(cluster_value):
                            # Map DataFrame index to sheet row: index 0 -> row 2, index 1 -> row 3, etc.
                            sheet_row = idx + 2
                            cell = ws.cell(row=sheet_row, column=cluster_col)
                            cell.value = int(cluster_value)
                            cluster_write_count += 1
                            self.logger.info(f"Wrote cluster {int(cluster_value)} to data point at sheet row {sheet_row} (DataFrame index {idx})")
                    self.logger.info(f"Total cluster assignments written to data rows: {cluster_write_count}")
                    
                    # Write centroid data to fixed rows
                    # Row 2-7 are reserved for cluster 0-5 centroid data
                    for cluster_num, centroid in cluster_centroids.items():
                        excel_row = int(cluster_num) + 2  # Row 2 for cluster 0, row 3 for cluster 1, etc.
                        # Always write the cluster number for identification
                        ws.cell(row=excel_row, column=cluster_col, value=int(cluster_num))
                        # Write centroid coordinates
                        ws.cell(row=excel_row, column=centroid_x_col, value=round(centroid[0], 4))
                        ws.cell(row=excel_row, column=centroid_y_col, value=round(centroid[1], 4))
                        ws.cell(row=excel_row, column=centroid_z_col, value=round(centroid[2], 4))
                        self.logger.info(f"Updated Excel row {excel_row} with cluster {int(cluster_num)} and centroid coordinates")
                    
                    # Save the workbook
                    wb.save(self.file_path)
                    self.logger.info("Excel file saved successfully")
                    
                    # Display success message
                    cluster_counts = valid_clusters.value_counts().to_dict()
                    cluster_info = "\\n".join(f"Cluster {k}: {v} points" for k, v in sorted(cluster_counts.items()))
                    
                    success_msg = (
                        f"Clusters and centroid coordinates saved for rows {start}-{end}!\\n\\n"
                        f"Cluster summary:\\n{cluster_info}\\n\\n"
                        f"Excel file has been updated with:\\n"
                        f"- Cluster assignments\\n"
                        f"- Centroid_X, Centroid_Y, Centroid_Z coordinates\\n\\n"
                        f"NEXT STEP: You can now calculate ΔE values by clicking the 'Calculate' button in the ΔE CIE2000 panel."
                    )
                    
                    messagebox.showinfo("Clusters Saved", success_msg)
                    return True
                else:
                    messagebox.showwarning("Warning", 
                        f"No cluster assignments found for rows {start}-{end}")
                    return False
                    
            finally:
                # Clean up backup if save was successful
                try:
                    if os.path.exists(backup_path):
                        os.remove(backup_path)
                except Exception:
                    self.logger.warning(f"Could not remove backup file: {backup_path}")
                    
        except Exception as e:
            self.logger.error(f"Error saving to Excel: {str(e)}")
            messagebox.showerror("Error", 
                f"Failed to save cluster assignments to Excel file.\\n\\n"
                f"Error: {str(e)}\\n\\n"
                f"Please check that the file isn't open in another program.")
            raise
    
    def _save_to_ods(self, start: int, end: int, row_indices) -> bool:
        """Save cluster assignments to an ODS (.ods) file using direct XML manipulation.
        
        Uses zipfile + lxml to modify content.xml directly, preserving all formatting
        and data validation.
        """
        import zipfile
        from lxml import etree
        
        # ODS namespaces
        NS = {
            'office': 'urn:oasis:names:tc:opendocument:xmlns:office:1.0',
            'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
            'text': 'urn:oasis:names:tc:opendocument:xmlns:text:1.0',
        }
        OFFICE_NS = '{urn:oasis:names:tc:opendocument:xmlns:office:1.0}'
        TABLE_NS = '{urn:oasis:names:tc:opendocument:xmlns:table:1.0}'
        TEXT_NS = '{urn:oasis:names:tc:opendocument:xmlns:text:1.0}'
        
        lockfile = None
        lock_path = f"{self.file_path}.lock"
        
        def set_cell_value(cell, value):
            """Set a cell's value while preserving its structure."""
            # Remove old text content
            for p in cell.findall(f'{TEXT_NS}p'):
                cell.remove(p)
            # Set value attributes
            cell.set(f'{OFFICE_NS}value-type', 'float')
            cell.set(f'{OFFICE_NS}value', str(value))
            # Add text for display
            p = etree.SubElement(cell, f'{TEXT_NS}p')
            p.text = str(value)
        
        def get_cell_at_column(row, col_idx):
            """Get the cell at a specific column, splitting repeated cells if needed.
            
            If the target column is within a repeated cell span, splits the repeated
            cell into individual cells so we can modify just one column.
            """
            cells = list(row.findall(f'{TABLE_NS}table-cell'))
            current_col = 0
            
            for cell in cells:
                repeat = cell.get(f'{TABLE_NS}number-columns-repeated')
                repeat_count = int(repeat) if repeat else 1
                
                if current_col <= col_idx < current_col + repeat_count:
                    # Found the cell containing our target column
                    if repeat_count == 1:
                        return cell
                    
                    # Need to split the repeated cell
                    offset = col_idx - current_col
                    
                    # Remove repeat attribute from original
                    if f'{TABLE_NS}number-columns-repeated' in cell.attrib:
                        del cell.attrib[f'{TABLE_NS}number-columns-repeated']
                    
                    # Get cell's position in row
                    cell_pos = list(row).index(cell)
                    
                    # Copy style if present
                    style = cell.get(f'{TABLE_NS}style-name')
                    
                    # Create cells BEFORE target (if offset > 0)
                    if offset > 0:
                        before_cell = etree.Element(f'{TABLE_NS}table-cell')
                        if style:
                            before_cell.set(f'{TABLE_NS}style-name', style)
                        if offset > 1:
                            before_cell.set(f'{TABLE_NS}number-columns-repeated', str(offset))
                        row.insert(cell_pos, before_cell)
                        cell_pos += 1  # Original cell shifted right
                    
                    # Create cells AFTER target (if any remain)
                    after_count = repeat_count - offset - 1
                    if after_count > 0:
                        after_cell = etree.Element(f'{TABLE_NS}table-cell')
                        if style:
                            after_cell.set(f'{TABLE_NS}style-name', style)
                        if after_count > 1:
                            after_cell.set(f'{TABLE_NS}number-columns-repeated', str(after_count))
                        row.insert(cell_pos + 1, after_cell)
                    
                    return cell
                
                current_col += repeat_count
            
            return None
        
        try:
            # Inform user of the process
            msg = (f"Saving cluster assignments for rows {start}-{end}:\\n\\n"
                  "1. Your original .ods file will be updated in-place\\n"
                  "2. Formatting and data validation will be preserved\\n\\n"
                  "Continue?")
            
            if not messagebox.askokcancel("Save Clusters", msg):
                return False
            
            # Try to acquire the lock with timeout
            max_attempts = 3
            attempt = 0
            lock_acquired = False
            
            while attempt < max_attempts and not lock_acquired:
                try:
                    lockfile = open(lock_path, 'w+')
                    fcntl.flock(lockfile, fcntl.LOCK_EX | fcntl.LOCK_NB)
                    lock_acquired = True
                    self.logger.info("Lock acquired successfully")
                except IOError as e:
                    if e.errno == errno.EACCES or e.errno == errno.EAGAIN:
                        attempt += 1
                        self.logger.warning(f"File is locked, attempt {attempt}/{max_attempts}")
                        
                        if attempt < max_attempts:
                            if not messagebox.askretrycancel("File Locked", 
                                f"The file appears to be in use by another program.\\n\\n"
                                f"Attempt {attempt} of {max_attempts}.\\n\\n"
                                "Would you like to try again?"):
                                if lockfile:
                                    lockfile.close()
                                return False
                            time.sleep(1)
                        else:
                            if lockfile:
                                lockfile.close()
                            messagebox.showerror("File Locked", 
                                "The file is locked by another program and cannot be accessed.\\n\\n"
                                "Please close any applications that might be using this file and try again.")
                            return False
                    else:
                        if lockfile:
                            lockfile.close()
                        raise
            
            if not lock_acquired:
                messagebox.showerror("File Locked", 
                    "Could not acquire file lock after multiple attempts.")
                return False
            
            # Get cluster assignments from in-memory data
            clusters = self.data.iloc[row_indices]['Cluster']
            valid_clusters = clusters[clusters.notna()]
            
            if not valid_clusters.empty:
                try:
                    # Create backup
                    backup_path = f"{self.file_path}.bak"
                    self.logger.info(f"Creating backup at: {backup_path}")
                    shutil.copy2(self.file_path, backup_path)
                    
                    # Get column indices using pandas
                    self.logger.info("Reading column structure with pandas")
                    df = pd.read_excel(self.file_path, engine='odf', sheet_name=self.sheet_name)
                    
                    # Verify required columns
                    required_columns = ['Cluster', 'Centroid_X', 'Centroid_Y', 'Centroid_Z']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    if missing_columns:
                        raise ValueError(f"Required columns missing: {missing_columns}")
                    
                    cluster_col_idx = list(df.columns).index('Cluster')
                    centroid_x_idx = list(df.columns).index('Centroid_X')
                    centroid_y_idx = list(df.columns).index('Centroid_Y')
                    centroid_z_idx = list(df.columns).index('Centroid_Z')
                    
                    self.logger.info(f"Column indices: Cluster={cluster_col_idx}, Centroid_X={centroid_x_idx}")
                    
                    # Calculate centroids
                    cluster_centroids = self._calculate_centroids()
                    
                    # Read ODS as ZIP and parse content.xml
                    self.logger.info("Reading ODS file as ZIP archive")
                    
                    # Read all files from the ODS archive
                    ods_files = {}
                    with zipfile.ZipFile(self.file_path, 'r') as zf:
                        for name in zf.namelist():
                            ods_files[name] = zf.read(name)
                    
                    # Parse content.xml
                    content_xml = ods_files['content.xml']
                    tree = etree.fromstring(content_xml)
                    
                    # Find the correct table (sheet)
                    tables = tree.xpath('//table:table', namespaces=NS)
                    if not tables:
                        raise ValueError("No tables found in ODS file")
                    
                    # Use specified sheet or default to first
                    table = None
                    if self.sheet_name:
                        for t in tables:
                            table_name = t.get(f'{TABLE_NS}name')
                            if table_name == self.sheet_name:
                                table = t
                                self.logger.info(f"Using specified ODS sheet: {self.sheet_name}")
                                break
                        if table is None:
                            self.logger.warning(f"Sheet '{self.sheet_name}' not found, using first sheet")
                            table = tables[0]
                    else:
                        table = tables[0]
                    
                    # Get all rows (row 0 is header, row 1+ is data)
                    rows = table.findall(f'{TABLE_NS}table-row')
                    self.logger.info(f"Found {len(rows)} rows in table")
                    
                    # Update cluster assignments in data rows
                    successful_writes = 0
                    for i, idx in enumerate(row_indices):
                        cluster_value = clusters.iloc[i]
                        if pd.notna(cluster_value):
                            # ODS row index = DataFrame index + 1 (for header)
                            ods_row_idx = idx + 1
                            if ods_row_idx < len(rows):
                                row = rows[ods_row_idx]
                                cell = get_cell_at_column(row, cluster_col_idx)
                                if cell is not None:
                                    set_cell_value(cell, int(cluster_value))
                                    successful_writes += 1
                                    self.logger.debug(f"✓ Updated ODS row {ods_row_idx} with cluster {int(cluster_value)}")
                                else:
                                    self.logger.warning(f"Could not find cell at row {ods_row_idx}, col {cluster_col_idx}")
                    
                    self.logger.info(f"Cluster assignment update summary: {successful_writes} rows updated")
                    
                    # Update centroid rows (rows 1, 2, 3 in ODS for clusters 0, 1, 2)
                    for cluster_num, centroid in cluster_centroids.items():
                        ods_row_idx = int(cluster_num) + 1  # +1 for header
                        if ods_row_idx < len(rows):
                            row = rows[ods_row_idx]
                            
                            # Set cluster number
                            cell = get_cell_at_column(row, cluster_col_idx)
                            if cell is not None:
                                set_cell_value(cell, int(cluster_num))
                            
                            # Set centroid coordinates
                            cell_x = get_cell_at_column(row, centroid_x_idx)
                            if cell_x is not None:
                                set_cell_value(cell_x, round(centroid[0], 4))
                            
                            cell_y = get_cell_at_column(row, centroid_y_idx)
                            if cell_y is not None:
                                set_cell_value(cell_y, round(centroid[1], 4))
                            
                            cell_z = get_cell_at_column(row, centroid_z_idx)
                            if cell_z is not None:
                                set_cell_value(cell_z, round(centroid[2], 4))
                            
                            self.logger.info(f"Updated centroid for cluster {cluster_num} at ODS row {ods_row_idx}")
                    
                    # Serialize modified XML and repackage ODS
                    self.logger.info("Repackaging ODS file")
                    modified_content = etree.tostring(tree, xml_declaration=True, encoding='UTF-8')
                    ods_files['content.xml'] = modified_content
                    
                    # Write new ODS file (preserving all other files like styles.xml)
                    temp_path = f"{self.file_path}.tmp"
                    with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                        for name, data in ods_files.items():
                            zf.writestr(name, data)
                    
                    # Replace original with updated file
                    os.replace(temp_path, self.file_path)
                    self.logger.info("File saved successfully")
                    
                    # Clean up backup
                    try:
                        if os.path.exists(backup_path):
                            os.remove(backup_path)
                    except Exception:
                        self.logger.warning(f"Could not remove backup file: {backup_path}")
                    
                    # Display success message
                    cluster_counts = valid_clusters.value_counts().to_dict()
                    cluster_info = "\\n".join(f"Cluster {k}: {v} points" for k, v in sorted(cluster_counts.items()))
                    
                    success_msg = (
                        f"Clusters and centroid coordinates saved for rows {start}-{end}!\\n\\n"
                        f"Cluster summary:\\n{cluster_info}\\n\\n"
                        f"Original .ods file has been updated with:\\n"
                        f"- Cluster assignments\\n"
                        f"- Centroid_X, Centroid_Y, Centroid_Z coordinates\\n\\n"
                        f"Formatting and data validation preserved.\\n\\n"
                        f"NEXT STEP: You can now calculate ΔE values by clicking the 'Calculate' button in the ΔE CIE2000 panel."
                    )
                    
                    messagebox.showinfo("Clusters Saved", success_msg)
                    return True
                    
                except Exception as e:
                    self.logger.error(f"Error updating spreadsheet: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    raise
            else:
                messagebox.showwarning("Warning", 
                    f"No cluster assignments found for rows {start}-{end}")
                return False
                
        except Exception as e:
            messagebox.showerror("Error", 
                "Failed to save cluster assignments.\\n\\n"
                "Please check file permissions and make sure the file isn't open in another program.")
            raise
        finally:
            if lockfile:
                self.logger.info("Releasing file lock")
                fcntl.flock(lockfile, fcntl.LOCK_UN)
                lockfile.close()
                try:
                    os.remove(lock_path)
                    self.logger.info("Lock file removed")
                except Exception as e:
                    self.logger.warning(f"Could not remove lock file: {str(e)}")
    
    def _calculate_centroids(self):
        """Calculate centroids for each cluster."""
        cluster_centroids = {}
        for cluster_num in self.data['Cluster'].dropna().unique():
            cluster_mask = self.data['Cluster'] == cluster_num
            centroid = [
                self.data.loc[cluster_mask, 'Xnorm'].mean(),
                self.data.loc[cluster_mask, 'Ynorm'].mean(),
                self.data.loc[cluster_mask, 'Znorm'].mean()
            ]
            cluster_centroids[int(cluster_num)] = centroid
            self.logger.info(f"Calculated centroid for cluster {int(cluster_num)}: {centroid}")
        return cluster_centroids
    
    def _verify_centroid_data(self, verify_doc, rows_to_verify, cluster_col_idx, centroid_col_indices):
        """Verify that centroid data was saved correctly."""
        verify_sheet = verify_doc.sheets[0]
        verification_count = min(5, len(rows_to_verify))
        self.logger.info(f"Verifying {verification_count} sample updates")
        
        for update in rows_to_verify[:verification_count]:
            cluster_value = update['cluster']
            row_idx = int(cluster_value) + 1
            centroid = update['centroid']
            
            # Verify cluster value
            cluster_cell = verify_sheet[row_idx, cluster_col_idx]
            if cluster_cell.value != cluster_value:
                self.logger.error(f"Cluster value mismatch at row {row_idx}")
                raise ValueError(f"Cluster save verification failed at row {row_idx}")
            
            # Verify centroid coordinates
            if centroid is not None:
                centroid_x_cell = verify_sheet[row_idx, centroid_col_indices['Centroid_X']]
                centroid_y_cell = verify_sheet[row_idx, centroid_col_indices['Centroid_Y']]
                centroid_z_cell = verify_sheet[row_idx, centroid_col_indices['Centroid_Z']]
                
                tolerance = 0.0001
                x_diff = abs(float(centroid_x_cell.value) - centroid[0])
                y_diff = abs(float(centroid_y_cell.value) - centroid[1])
                z_diff = abs(float(centroid_z_cell.value) - centroid[2])
                
                if x_diff > tolerance or y_diff > tolerance or z_diff > tolerance:
                    self.logger.error(f"Centroid coordinate mismatch at row {row_idx}")
                    raise ValueError(f"Centroid save verification failed at row {row_idx}")
        
        centroid_cols = ", ".join([f"{col}:{idx}" for col, idx in centroid_col_indices.items()])
        self.logger.info(f"Centroid columns successfully verified: {centroid_cols}")
