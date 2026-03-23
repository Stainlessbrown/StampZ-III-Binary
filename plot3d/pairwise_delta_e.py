#!/usr/bin/env python3
"""
Pairwise ΔE Matrix — Shade Consistency Tool for Plot_3D.

Computes an N×N pairwise ΔE CIE2000 matrix for a selected set of colour
data points and displays it in a colour-coded viewer with export to ODS/XLSX.
"""

import os
import math
import logging
import tkinter as tk
from tkinter import ttk, messagebox
from typing import List, Tuple, Dict, Optional, Any

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# ΔE CIE2000 implementation  (self-contained — mirrors DeltaEManager)
# ---------------------------------------------------------------------------

def _delta_e_2000(lab1: Tuple[float, float, float],
                  lab2: Tuple[float, float, float]) -> float:
    """Calculate ΔE CIE2000 between two L*a*b* colours."""
    L1, a1, b1 = lab1
    L2, a2, b2 = lab2

    kL = kC = kH = 1.0

    C1 = math.sqrt(a1 ** 2 + b1 ** 2)
    C2 = math.sqrt(a2 ** 2 + b2 ** 2)
    Cab = (C1 + C2) / 2
    G = 0.5 * (1 - math.sqrt(Cab ** 7 / (Cab ** 7 + 25 ** 7)))

    a1p = (1 + G) * a1
    a2p = (1 + G) * a2
    C1p = math.sqrt(a1p ** 2 + b1 ** 2)
    C2p = math.sqrt(a2p ** 2 + b2 ** 2)

    def _h_prime(ap, b):
        if ap == 0 and b == 0:
            return 0
        h = math.degrees(math.atan2(b, ap))
        return h + 360 if h < 0 else h

    h1p = _h_prime(a1p, b1)
    h2p = _h_prime(a2p, b2)

    deltaLp = L2 - L1
    deltaCp = C2p - C1p

    dhp = h2p - h1p
    if C1p * C2p == 0:
        deltaHp_angle = 0
    elif abs(dhp) <= 180:
        deltaHp_angle = dhp
    elif dhp > 180:
        deltaHp_angle = dhp - 360
    else:
        deltaHp_angle = dhp + 360
    deltaHp = 2 * math.sqrt(C1p * C2p) * math.sin(math.radians(deltaHp_angle / 2))

    Lp = (L1 + L2) / 2
    Cp = (C1p + C2p) / 2

    if C1p * C2p == 0:
        hp = h1p + h2p
    elif abs(h1p - h2p) <= 180:
        hp = (h1p + h2p) / 2
    elif h1p + h2p < 360:
        hp = (h1p + h2p + 360) / 2
    else:
        hp = (h1p + h2p - 360) / 2

    T = (1
         - 0.17 * math.cos(math.radians(hp - 30))
         + 0.24 * math.cos(math.radians(2 * hp))
         + 0.32 * math.cos(math.radians(3 * hp + 6))
         - 0.20 * math.cos(math.radians(4 * hp - 63)))

    dTheta = 30 * math.exp(-((hp - 275) / 25) ** 2)
    RC = 2 * math.sqrt(Cp ** 7 / (Cp ** 7 + 25 ** 7))
    RT = -math.sin(math.radians(2 * dTheta)) * RC

    SL = 1 + (0.015 * (Lp - 50) ** 2) / math.sqrt(20 + (Lp - 50) ** 2)
    SC = 1 + 0.045 * Cp
    SH = 1 + 0.015 * Cp * T

    return math.sqrt(
        (deltaLp / (kL * SL)) ** 2
        + (deltaCp / (kC * SC)) ** 2
        + (deltaHp / (kH * SH)) ** 2
        + RT * (deltaCp / (kC * SC)) * (deltaHp / (kH * SH))
    )


def _denormalize_lab(x: float, y: float, z: float) -> Tuple[float, float, float]:
    """Convert normalised 0-1 Plot_3D coordinates to standard L*a*b* ranges.

    Mapping (matches DeltaEManager / DeltaECalculator):
        L* = X × 100          (0-100)
        a* = Y × 255 - 128    (-128 to +127)
        b* = Z × 255 - 128    (-128 to +127)
    """
    return (x * 100, y * 255 - 128, z * 255 - 128)


def _rgb_to_lab(r: float, g: float, b: float) -> Tuple[float, float, float]:
    """Convert normalised sRGB (0-1) to L*a*b* via XYZ (D65)."""
    def srgb_to_linear(c):
        return c / 12.92 if c <= 0.04045 else ((c + 0.055) / 1.055) ** 2.4

    rl, gl, bl = srgb_to_linear(r), srgb_to_linear(g), srgb_to_linear(b)

    # sRGB → XYZ
    x = rl * 0.4124564 + gl * 0.3575761 + bl * 0.1804375
    y = rl * 0.2126729 + gl * 0.7151522 + bl * 0.0721750
    z = rl * 0.0193339 + gl * 0.1191920 + bl * 0.9503041

    # D65 white point
    x /= 0.95047
    y /= 1.00000
    z /= 1.08883

    def f(t):
        return t ** (1 / 3) if t > 0.008856 else (903.3 * t + 16) / 116

    fx, fy, fz = f(x), f(y), f(z)
    L = max(0.0, min(100.0, 116 * fy - 16))
    a = 500 * (fx - fy)
    b_star = 200 * (fy - fz)
    return L, a, b_star


# ---------------------------------------------------------------------------
# PairwiseDeltaEManager
# ---------------------------------------------------------------------------

class PairwiseDeltaEManager:
    """Compute and display the pairwise ΔE matrix for a selection of points."""

    def __init__(self, logger=None, color_space: str = 'LAB'):
        if logger is None:
            self.logger = logging.getLogger("PairwiseDeltaE")
            if not self.logger.handlers:
                h = logging.StreamHandler()
                h.setFormatter(logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s'))
                self.logger.addHandler(h)
            self.logger.setLevel(logging.INFO)
            self.logger.propagate = False
        else:
            self.logger = logger

        self.color_space = color_space.upper() if color_space else 'LAB'
        self.file_path: Optional[str] = None
        self.sheet_name: Optional[str] = None
        self.data: Optional[pd.DataFrame] = None

        # GUI widgets (created by create_gui)
        self.frame = None
        self.start_row_spin = None
        self.end_row_spin = None
        self.threshold_spin = None
        self.cluster_var = None

    # -- data plumbing (mirrors DeltaEManager) ----------------------------

    def set_file_path(self, path: str):
        self.file_path = os.path.abspath(path) if path else None

    def set_sheet_name(self, name: str):
        self.sheet_name = name

    def load_data(self, df: pd.DataFrame):
        if df is not None:
            self.data = df.copy()

    # -- row-index helpers (handles ODS / XLSX / realtime offset) ----------

    def _get_row_indices(self, start: int, end: int) -> List[int]:
        """Convert 1-based display rows → 0-based DataFrame indices.

        Handles the realtime worksheet ``_original_sheet_row`` mapping when
        present, otherwise uses the standard ``row - 2`` conversion (row 2
        in the UI = index 0).
        """
        if self.data is None:
            return []

        if '_original_sheet_row' in self.data.columns:
            # Realtime mode: display row N → sheet row N-1
            sheet_start = start - 1
            sheet_end = end - 1
            return [
                idx for idx, row in self.data.iterrows()
                if pd.notna(row.get('_original_sheet_row'))
                and sheet_start <= row['_original_sheet_row'] <= sheet_end
            ]

        # File mode
        zero_start = max(start - 2, 0)
        zero_end = min(end - 1, len(self.data) - 1)
        return list(range(zero_start, zero_end + 1))

    # -- core computation --------------------------------------------------

    def _point_to_lab(self, row: pd.Series) -> Optional[Tuple[float, float, float]]:
        """Extract L*a*b* from a data row, handling LAB / RGB / CMY spaces."""
        x = row.get('Xnorm')
        y = row.get('Ynorm')
        z = row.get('Znorm')
        if any(pd.isna(v) for v in (x, y, z)):
            return None

        if self.color_space == 'LAB':
            return _denormalize_lab(x, y, z)
        elif self.color_space == 'CMY':
            return _rgb_to_lab(1 - x, 1 - y, 1 - z)
        else:  # RGB
            return _rgb_to_lab(x, y, z)

    def compute_matrix(self, start_row: int, end_row: int,
                       cluster_filter: Optional[int] = None
                       ) -> Tuple[List[str], List[List[float]], Dict[str, Any]]:
        """Compute the pairwise ΔE matrix.

        Returns:
            data_ids  – list of DataID labels (length N)
            matrix    – N×N list-of-lists of ΔE values
            stats     – dict with max_de, mean_de, pairs_above, total_pairs, pct_within
        """
        if self.data is None:
            raise ValueError("No data loaded.")

        # Reload from file if available to ensure freshness
        if self.file_path and os.path.exists(self.file_path):
            engine = 'odf' if self.file_path.endswith('.ods') else 'openpyxl'
            self.data = pd.read_excel(self.file_path, engine=engine,
                                      sheet_name=self.sheet_name or 0)
            self.logger.info(f"Reloaded data from {self.file_path}")

        indices = self._get_row_indices(start_row, end_row)
        if not indices:
            raise ValueError(f"No data rows found for range {start_row}-{end_row}")

        subset = self.data.iloc[indices]

        # Optional cluster filter
        if cluster_filter is not None and 'Cluster' in subset.columns:
            subset = subset[subset['Cluster'] == cluster_filter]
            if subset.empty:
                raise ValueError(f"No rows found for cluster {cluster_filter}")

        # Build parallel lists of DataIDs and Lab values
        data_ids: List[str] = []
        labs: List[Tuple[float, float, float]] = []

        for _, row in subset.iterrows():
            lab = self._point_to_lab(row)
            if lab is None:
                continue
            did = str(row.get('DataID', f'Row {_ + 2}'))
            data_ids.append(did)
            labs.append(lab)

        n = len(labs)
        if n < 2:
            raise ValueError(f"Need at least 2 valid points; found {n}.")

        self.logger.info(f"Computing {n}×{n} pairwise matrix ({n*(n-1)//2} unique pairs)")

        matrix = [[0.0] * n for _ in range(n)]
        all_de: List[float] = []

        for i in range(n):
            for j in range(i + 1, n):
                de = round(_delta_e_2000(labs[i], labs[j]), 2)
                matrix[i][j] = de
                matrix[j][i] = de
                all_de.append(de)

        # Threshold for "same shade" — caller can override via GUI
        threshold = 2.3
        if hasattr(self, '_gui_threshold'):
            threshold = self._gui_threshold

        pairs_above = sum(1 for d in all_de if d > threshold)
        stats = {
            'max_de': max(all_de) if all_de else 0.0,
            'mean_de': round(sum(all_de) / len(all_de), 2) if all_de else 0.0,
            'pairs_above': pairs_above,
            'total_pairs': len(all_de),
            'pct_within': round((1 - pairs_above / len(all_de)) * 100, 1) if all_de else 100.0,
            'threshold': threshold,
        }

        self.logger.info(f"Matrix complete — max={stats['max_de']}, mean={stats['mean_de']}, "
                         f"{stats['pct_within']}% within ΔE {threshold}")
        return data_ids, matrix, stats

    # -- GUI ---------------------------------------------------------------

    def create_gui(self, parent) -> tk.Frame:
        """Create the compact control panel (mirrors DeltaEManager style)."""
        self.frame = tk.Frame(parent, bg='white')
        self.frame.grid_columnconfigure(0, weight=1)

        # Separator label
        tk.Label(self.frame, text="── Pairwise ΔE ──", font=("Arial", 9, "bold"),
                 bg='white', fg='#555').pack(fill=tk.X, pady=(8, 2))

        row_frame = tk.Frame(self.frame, bg='white')
        row_frame.pack(fill=tk.X, padx=5, pady=4)

        tk.Label(row_frame, text="Rows:", font=("Arial", 9), bg='white').pack(side=tk.LEFT, padx=2)
        self.start_row_spin = tk.Spinbox(row_frame, from_=2, to=999, width=4,
                                         justify='center', font=("Arial", 9))
        self.start_row_spin.delete(0, tk.END)
        self.start_row_spin.insert(0, "2")
        self.start_row_spin.pack(side=tk.LEFT, padx=1)

        tk.Label(row_frame, text="to", font=("Arial", 9), bg='white').pack(side=tk.LEFT, padx=2)
        self.end_row_spin = tk.Spinbox(row_frame, from_=2, to=999, width=4,
                                       justify='center', font=("Arial", 9))
        self.end_row_spin.delete(0, tk.END)
        self.end_row_spin.insert(0, "999")
        self.end_row_spin.pack(side=tk.LEFT, padx=1)

        tk.Label(row_frame, text="ΔE≤", font=("Arial", 9), bg='white').pack(side=tk.LEFT, padx=(6, 1))
        self.threshold_spin = tk.Spinbox(row_frame, from_=0.5, to=10.0, increment=0.5,
                                         width=4, justify='center', font=("Arial", 9),
                                         format="%.1f")
        self.threshold_spin.delete(0, tk.END)
        self.threshold_spin.insert(0, "2.3")
        self.threshold_spin.pack(side=tk.LEFT, padx=1)

        btn_frame = tk.Frame(self.frame, bg='white')
        btn_frame.pack(fill=tk.X, padx=5, pady=(0, 4))

        # Optional cluster filter
        tk.Label(btn_frame, text="Cluster:", font=("Arial", 9), bg='white').pack(side=tk.LEFT, padx=2)
        self.cluster_var = tk.StringVar(value="All")
        self.cluster_combo = ttk.Combobox(btn_frame, textvariable=self.cluster_var,
                                          values=["All"], width=5, state='readonly',
                                          font=("Arial", 9))
        self.cluster_combo.pack(side=tk.LEFT, padx=2)

        calc_btn = tk.Button(btn_frame, text="Matrix", command=self._on_calculate,
                             font=("Arial", 9), bg="#d0e8ff")
        calc_btn.pack(side=tk.LEFT, padx=4)

        help_btn = tk.Button(btn_frame, text="?", command=self._show_help,
                             font=("Arial", 9), width=2, bg="lightblue")
        help_btn.pack(side=tk.LEFT, padx=2)

        # Populate cluster dropdown from data
        self._refresh_cluster_list()

        return self.frame

    def _refresh_cluster_list(self):
        """Populate the cluster filter dropdown."""
        values = ["All"]
        if self.data is not None and 'Cluster' in self.data.columns:
            clusters = sorted(self.data['Cluster'].dropna().unique())
            values += [str(int(c)) for c in clusters]
        if hasattr(self, 'cluster_combo'):
            self.cluster_combo['values'] = values

    def _on_calculate(self):
        """Handle the Matrix button click."""
        try:
            start = int(self.start_row_spin.get())
            end = int(self.end_row_spin.get())
            threshold = float(self.threshold_spin.get())
        except ValueError:
            messagebox.showerror("Invalid Input", "Row and threshold values must be numbers.")
            return

        self._gui_threshold = threshold

        cluster_filter = None
        cv = self.cluster_var.get()
        if cv != "All":
            try:
                cluster_filter = int(cv)
            except ValueError:
                pass

        try:
            data_ids, matrix, stats = self.compute_matrix(start, end, cluster_filter)
            # Open the viewer window
            PairwiseDeltaEViewer(
                parent=self.frame.winfo_toplevel(),
                data_ids=data_ids,
                matrix=matrix,
                stats=stats,
                file_path=self.file_path,
                sheet_name=self.sheet_name,
            )
        except Exception as e:
            self.logger.error(f"Pairwise ΔE error: {e}")
            messagebox.showerror("Pairwise ΔE Error", str(e))

    def _show_help(self):
        messagebox.showinfo("Pairwise ΔE Help",
            "Shade Consistency — Pairwise ΔE Matrix\n\n"
            "Computes ΔE CIE2000 between every pair of\n"
            "points in the selected row range.\n\n"
            "• Rows: data row range (row 2 = first data row)\n"
            "• ΔE≤: threshold for 'same shade' highlighting\n"
            "• Cluster: optionally limit to one cluster\n\n"
            "The matrix window colour-codes cells:\n"
            "  Green  — ΔE ≤ 1.0  (imperceptible)\n"
            "  Yellow — ΔE 1.0–threshold (same shade)\n"
            "  Red    — ΔE > threshold (different shade)\n\n"
            "Use 'Export' to write the matrix as a new sheet\n"
            "in the source spreadsheet file."
        )


# ---------------------------------------------------------------------------
# PairwiseDeltaEViewer — matrix display window
# ---------------------------------------------------------------------------

class PairwiseDeltaEViewer:
    """Toplevel window showing the colour-coded pairwise ΔE matrix."""

    CELL_W = 72
    CELL_H = 30
    HEADER_W = 110
    HEADER_H = 30

    def __init__(self, parent, data_ids: List[str], matrix: List[List[float]],
                 stats: Dict[str, Any], file_path: Optional[str] = None,
                 sheet_name: Optional[str] = None):
        self.data_ids = data_ids
        self.matrix = matrix
        self.stats = stats
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.n = len(data_ids)

        self.win = tk.Toplevel(parent)
        self.win.title("Pairwise ΔE Matrix — Shade Consistency")
        # Size the window to fit or cap at 900×700
        w = min(self.HEADER_W + self.n * self.CELL_W + 40, 1200)
        h = min(self.HEADER_H + self.n * self.CELL_H + 120, 800)
        self.win.geometry(f"{w}x{h}")
        self.win.minsize(400, 300)

        self._build_ui()

    def _build_ui(self):
        s = self.stats
        threshold = s.get('threshold', 2.3)

        # Summary bar
        summary_frame = tk.Frame(self.win, bg='#f0f0f0')
        summary_frame.pack(fill=tk.X, padx=5, pady=5)

        summary_text = (
            f"Max: {s['max_de']:.2f}  |  Mean: {s['mean_de']:.2f}  |  "
            f"Pairs > {threshold}: {s['pairs_above']}/{s['total_pairs']}  |  "
            f"Within threshold: {s['pct_within']}%"
        )
        tk.Label(summary_frame, text=summary_text, font=("Arial", 10, "bold"),
                 bg='#f0f0f0').pack(side=tk.LEFT, padx=10, pady=4)

        # Export button
        if self.file_path:
            tk.Button(summary_frame, text="Export to File", command=self._export,
                      font=("Arial", 9), bg="#d0ffd0").pack(side=tk.RIGHT, padx=10)

        # Scrollable canvas for the matrix
        canvas_frame = tk.Frame(self.win)
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.canvas = tk.Canvas(canvas_frame, bg='white')
        h_scroll = tk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        v_scroll = tk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.canvas.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)

        self.canvas.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        self._draw_matrix()

    def _cell_colour(self, de: float, is_diag: bool = False) -> str:
        threshold = self.stats.get('threshold', 2.3)
        if is_diag:
            return '#e0e0e0'
        if de <= 1.0:
            return '#c8f7c8'    # green — imperceptible
        if de <= threshold:
            return '#fff4b0'    # yellow — same shade
        return '#ffb0b0'        # red — different shade

    def _draw_matrix(self):
        cw, ch = self.CELL_W, self.CELL_H
        hw, hh = self.HEADER_W, self.HEADER_H
        n = self.n

        # Column headers (DataIDs rotated is hard in Canvas; use abbreviated text)
        for j in range(n):
            x = hw + j * cw
            label = self.data_ids[j]
            if len(label) > 8:
                label = label[:7] + "…"
            self.canvas.create_text(x + cw // 2, hh // 2, text=label,
                                    font=("Arial", 11, "bold"), anchor='center')

        # Rows
        for i in range(n):
            y = hh + i * ch
            # Row header
            label = self.data_ids[i]
            if len(label) > 10:
                label = label[:9] + "…"
            self.canvas.create_text(hw - 4, y + ch // 2, text=label,
                                    font=("Arial", 11, "bold"), anchor='e')

            for j in range(n):
                x = hw + j * cw
                de = self.matrix[i][j]
                is_diag = (i == j)
                fill = self._cell_colour(de, is_diag)
                self.canvas.create_rectangle(x, y, x + cw, y + ch, fill=fill, outline='#cccccc')
                text = "—" if is_diag else f"{de:.2f}"
                self.canvas.create_text(x + cw // 2, y + ch // 2, text=text,
                                        font=("Arial", 11), anchor='center')

        total_w = hw + n * cw + 10
        total_h = hh + n * ch + 10
        self.canvas.configure(scrollregion=(0, 0, total_w, total_h))

    # -- export ------------------------------------------------------------

    def _export(self):
        """Write the matrix as a new sheet in the source file."""
        if not self.file_path or not os.path.exists(self.file_path):
            messagebox.showerror("Export Error", "Source file not found.")
            return

        try:
            if self.file_path.endswith('.xlsx'):
                self._export_xlsx()
            else:
                self._export_ods()
            messagebox.showinfo("Export Complete",
                                f"Pairwise ΔE matrix written as a new sheet\n"
                                f"in {os.path.basename(self.file_path)}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export matrix:\n\n{e}")

    def _export_xlsx(self):
        from openpyxl import load_workbook

        wb = load_workbook(self.file_path)
        # Remove existing sheet if present
        if 'Pairwise ΔE' in wb.sheetnames:
            del wb['Pairwise ΔE']
        ws = wb.create_sheet('Pairwise ΔE')

        n = self.n
        # Header row (row 1): blank + DataIDs
        for j in range(n):
            ws.cell(row=1, column=j + 2, value=self.data_ids[j])

        # Data rows
        for i in range(n):
            ws.cell(row=i + 2, column=1, value=self.data_ids[i])
            for j in range(n):
                ws.cell(row=i + 2, column=j + 2, value=self.matrix[i][j])

        # Summary below matrix
        summary_row = n + 4
        s = self.stats
        ws.cell(row=summary_row, column=1, value="Summary")
        ws.cell(row=summary_row + 1, column=1, value="Max ΔE")
        ws.cell(row=summary_row + 1, column=2, value=s['max_de'])
        ws.cell(row=summary_row + 2, column=1, value="Mean ΔE")
        ws.cell(row=summary_row + 2, column=2, value=s['mean_de'])
        ws.cell(row=summary_row + 3, column=1, value=f"Pairs > {s['threshold']}")
        ws.cell(row=summary_row + 3, column=2, value=f"{s['pairs_above']}/{s['total_pairs']}")
        ws.cell(row=summary_row + 4, column=1, value="% Within threshold")
        ws.cell(row=summary_row + 4, column=2, value=f"{s['pct_within']}%")

        wb.save(self.file_path)

    def _export_ods(self):
        import ezodf

        doc = ezodf.opendoc(self.file_path)

        # Remove existing sheet if present
        for i, sheet in enumerate(doc.sheets):
            if sheet.name == 'Pairwise ΔE':
                del doc.sheets[i]
                break

        n = self.n
        total_rows = n + 7  # data + header + spacing + summary
        total_cols = n + 2
        new_sheet = ezodf.Sheet('Pairwise ΔE', size=(total_rows, total_cols))
        doc.sheets += new_sheet

        # Header row (row 0): blank + DataIDs
        for j in range(n):
            new_sheet[0, j + 1].set_value(self.data_ids[j])

        # Data rows
        for i in range(n):
            new_sheet[i + 1, 0].set_value(self.data_ids[i])
            for j in range(n):
                new_sheet[i + 1, j + 1].set_value(self.matrix[i][j])

        # Summary below matrix
        sr = n + 3
        s = self.stats
        new_sheet[sr, 0].set_value("Summary")
        new_sheet[sr + 1, 0].set_value("Max ΔE")
        new_sheet[sr + 1, 1].set_value(s['max_de'])
        new_sheet[sr + 2, 0].set_value("Mean ΔE")
        new_sheet[sr + 2, 1].set_value(s['mean_de'])
        new_sheet[sr + 3, 0].set_value(f"Pairs > {s['threshold']}")
        new_sheet[sr + 3, 1].set_value(f"{s['pairs_above']}/{s['total_pairs']}")

        doc.save()
