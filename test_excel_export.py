#!/usr/bin/env python3
"""
Test Excel Export Specifically

Tests the Excel export functionality to ensure data is properly populated.
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from utils.rgb_cmy_color_analyzer import RGBCMYColorAnalyzer
from PIL import Image, ImageDraw


class MockCoordinateMarker:
    def __init__(self, x, y, width=20, height=20, shape='rectangle', anchor='center'):
        self.x = x
        self.y = y
        self.width = width
        self.height = height
        self.shape = shape
        self.anchor = anchor


def test_excel_export():
    """Test Excel export with data population."""
    print("📊 Testing Excel Export with Data Population")
    print("=" * 50)
    
    # Create test image with distinct colors
    test_image = Image.new('RGB', (600, 400), (255, 255, 255))
    draw = ImageDraw.Draw(test_image)
    
    # Create three distinct colored regions
    colors = [
        ((50, 50, 100, 100), (200, 50, 50)),   # Red-ish
        ((200, 50, 100, 100), (50, 200, 50)), # Green-ish 
        ((350, 50, 100, 100), (50, 50, 200)), # Blue-ish
    ]
    
    for (x, y, w, h), color in colors:
        draw.rectangle([x, y, x+w, y+h], fill=color)
    
    test_image_path = "test_excel_export.png"
    test_image.save(test_image_path)
    print(f"📸 Created test image: {test_image_path}")
    
    # Create markers
    coord_markers = [
        MockCoordinateMarker(100, 100, 50, 50),  # Red region
        MockCoordinateMarker(250, 100, 50, 50),  # Green region
        MockCoordinateMarker(400, 100, 50, 50),  # Blue region
    ]
    
    # Run analysis
    analyzer = RGBCMYColorAnalyzer()
    results = analyzer.analyze_image_rgb_cmy_from_canvas(
        test_image_path,
        "Excel_Export_Test",
        coord_markers
    )
    
    if not results:
        print("❌ Analysis failed!")
        return False
    
    print("✅ Analysis completed")
    print("📊 Sample Results:")
    for i, result in enumerate(results['results']):
        print(f"  Sample {i+1}: R={result['R_mean']:.1f}, G={result['G_mean']:.1f}, B={result['B_mean']:.1f}")
    
    # Test Excel export
    template_path = "/Users/stanbrown/Desktop/StampZ-III-Binary/data/templates/RGB-CMY Channel analysis.xlsx"
    output_path = "Test_Excel_Export_Result.xlsx"
    
    print(f"\n📤 Testing Excel export...")
    print(f"   Template: {template_path}")
    print(f"   Output: {output_path}")
    
    rgb_cmy_analyzer = results['analyzer']
    
    # Test the Excel export method directly
    success = rgb_cmy_analyzer._export_to_xlsx(template_path, output_path)
    
    if success:
        print(f"✅ Excel export successful!")
        print(f"📁 Created: {output_path}")
        
        # Try to read it back to verify
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            print(f"\n🔍 Verifying Excel content:")
            print(f"   B2 (Sample Set): {ws['B2'].value}")
            print(f"   B16 (Sample 1 B): {ws['B16'].value}")
            print(f"   E16 (Sample 1 G): {ws['E16'].value}")
            print(f"   H16 (Sample 1 R): {ws['H16'].value}")
            print(f"   B29 (Sample 1 C): {ws['B29'].value}")
            print(f"   E29 (Sample 1 Y): {ws['E29'].value}")
            print(f"   H29 (Sample 1 M): {ws['H29'].value}")
            
            # Check if formulas are working
            formula_cells = ['D16', 'G16', 'J16']  # Should be 1/SD² formulas
            for cell in formula_cells:
                cell_obj = ws[cell]
                print(f"   {cell} (formula): {cell_obj.value}")
                
        except Exception as e:
            print(f"❌ Error reading Excel file: {e}")
    else:
        print(f"❌ Excel export failed!")
    
    return success


if __name__ == "__main__":
    test_excel_export()